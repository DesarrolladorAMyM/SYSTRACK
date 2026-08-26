"""
views.py — SYSTRAKER
APIs REST para el dashboard de inventario.
Todas las respuestas son JSON para consumo del frontend.
"""

import json
import re
from datetime import date
from django.http import JsonResponse, HttpResponse
from requerimientos.models import Usuario,CentroOperacion,Cargo,TipoUsuario,Requerimiento

from django.views.decorators.http import require_GET,require_POST
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.shortcuts import render, get_object_or_404
from django.db import transaction
from django.db.models import Count, Q ,F
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .models import (
    TipoDispositivo, Estado, Marca, Propietario, Departamento,
    Municipio, TipoDocumento, CentroOperaciones, Antivirus, Procesador,
    SistemaOperativo, LicenciaOffice, Opciones, Almacenamiento,
    TipoNovedad, Operador, Dispositivo, CaracteristicaPC,
    CaracteristicaMovil, CaracteristicaPantalla, CaracteristicaImpresora,
    CaracteristicaPeriferico, CaracteristicaLicencia,
    HistorialEquipo, Colaborador,
    AsignacionColaborador, Acta, ActaDispositivo, CentroCosto, TipoImpresora,
    RAM, TipoDisco ,CaracteristicasVideoBeam,TipoActa, NotificacionBellLeida,
    ItemChecklist, ChecklistDispositivo, RespuestaChecklist,
    TipoNovedadGeneral, CampoNovedadGeneral, NovedadGeneral, RespuestaNovedadGeneral
)
import base64
import os
from zoneinfo import ZoneInfo
from django.conf import settings

# El servidor guarda las fechas en UTC (TIME_ZONE='UTC' en settings), así que
# hay que convertir a hora de Colombia antes de mostrarlas al usuario.
COLOMBIA_TZ = ZoneInfo('America/Bogota')


def _fecha_local_str(dt, fmt='%d/%m/%Y %H:%M'):
    if not dt:
        return ''
    return timezone.localtime(dt, COLOMBIA_TZ).strftime(fmt)


#  UTILIDADES — Generación PDF y envío de correo del Acta

import smtplib
import threading
from email.mime.multipart import MIMEMultipart
from email.mime.text       import MIMEText
from email.mime.base       import MIMEBase
from email                 import encoders
from xhtml2pdf import pisa
from io import BytesIO

# Estados de Dispositivo que se consideran "inactivos": estos equipos se
# excluyen del Inventario activo y se listan en la pantalla de Inactivos.
ESTADOS_INACTIVOS = ('ELIMINADO', 'OBSOLETO', 'DEVUELTO')

# Fecha desde la que la campanita avisa de dispositivos sin checklist (ver
# api_notificaciones_bell). Fecha fija a propósito: los dispositivos creados
# antes de que el checklist existiera como función no cuentan como pendientes.
CHECKLIST_NOTIF_DESDE = date(2026, 8, 26)


# VISTA PRINCIPAL  Renderiza el dashboard HTML
@login_required(login_url='login')
def dashboard(request):
   
    return render(request, 'dashboard/dashboard.html')



# UTILIDADES

def _json_ok(data):
    return JsonResponse({'ok': True, 'data': data})

def _json_err(msg, status=400):
    return JsonResponse({'ok': False, 'error': msg}, status=status)


# ═══════════════════════════════════════════════════
#  CATÁLOGOS — GET para poblar los <select> del HTML
# ═══════════════════════════════════════════════════
@login_required(login_url='login')
@require_http_methods(['GET'])
def api_catalogos(request): 
    """
    Retorna todos los catálogos necesarios para poblar los
    <select> del formulario de inventario y demás modales.
    Un solo endpoint para reducir round-trips al cargar la página.
    """
    data = {
        # Catálogos básicos
        'tipos_dispositivo': list(
            TipoDispositivo.objects.filter(g200_estado=True)
            .values('g200_id', 'g200_tipo_dispositivo')
        ),
        'estados': list(
            Estado.objects.filter(g201_estado=True)
            .values('g201_id', 'g201_descripcion')
        ),
        'marcas': list(
            Marca.objects.filter(g202_estado=True)
            .values('g202_id', 'g202_marca')
        ),
        'propietarios': list(
            Propietario.objects.filter(g203_estado=True)
            .values('g203_id', 'g203_propietario', 'g203_documento')
        ),
        'departamentos': list(
            Departamento.objects.filter(g204_estado=True)
            .values('g204_id', 'g204_departamento')
        ),
        'centros_costo': list(
            CentroCosto.objects.filter(g228_estado=True)
            .values('g228_id', 'g228_nombre')
        ),
        'centros_operaciones': list(
            CentroOperaciones.objects.filter(g207_estado=True)
            .values('g207_id', 'g207_co', 'g207_descripcion_co')
        ),
        'tipos_novedad': list(
            TipoNovedad.objects.filter(g220_estado=True)
            .values('g220_id', 'g220_novedad')
        ),
        'tipos_documento': list(
            TipoDocumento.objects.filter(g206_estado=True)
            .values('g206_id', 'g206_tipo_documento')
        ),

        # Catálogos de características PC
        'antivirus': list(
            Antivirus.objects.filter(g208_estado=True)
            .values('g208_id', 'g208_antivirus')
        ),
        'procesadores': list(
            Procesador.objects.filter(g209_estado=True)
            .values('g209_id', 'g209_procesador')
        ),
        'sistemas_operativos': list(
            SistemaOperativo.objects.filter(g210_estado=True)
            .values('g210_id', 'g210_so')
        ),
        'licencias_office': list(
            LicenciaOffice.objects.filter(g211_estado=True)
            .values('g211_id', 'g211_office')
        ),
        'opciones': list(
            Opciones.objects.filter(g218_estado=True)
            .values('g218_id', 'g218_opciones')
        ),
        'almacenamientos': list(
            Almacenamiento.objects.filter(g219_estado=True)
            .values('g219_id', 'g219_almacenamiento')
        ),

        # Catálogos móviles
        'operadores': list(
            Operador.objects.filter(g221_estado=True)
            .values('g221_id', 'g221_operador')
        ),
        'tipos_impresora':list(
          TipoImpresora.objects.filter(g229_estado=True).
          values('g229_id','g229_tipo_impresora')  
        ),
        'rams': list(
            RAM.objects.filter(g230_estado=True)
            .values('g230_id', 'g230_ram')
        ),
        'tipos_disco': list(
            TipoDisco.objects.filter(g231_estado=True)
            .values('g231_id', 'g231_tipo_disco')
        ),
        'tipos_acta': list(
            TipoActa.objects.filter(g233_estado=True)
            .values('g233_id', 'g233_tipo_acta')
        ),
        
    }
    
    
    return _json_ok(data)
@login_required(login_url='login')
@require_http_methods(['GET'])
def api_municipios_por_dpto(request, dpto_id):
    """Retorna los municipios de un departamento específico."""
    municipios = list(
        Municipio.objects.filter(
            g205_departamento_id=dpto_id,
            g205_estado=True
        ).values('g205_id', 'g205_municipio')
    )
    return _json_ok(municipios)


# ═══════════════════════════════════════════════════
#  INVENTARIO — DISPOSITIVOS ACTIVOS
# ═══════════════════════════════════════════════════
@login_required(login_url='login')
@require_http_methods(['GET'])
def api_dispositivos(request):
    """
    Lista de dispositivos activos con filtros opcionales:
      ?q=       búsqueda libre (serial, propietario, marca)
      ?tipo=    id del TipoDispositivo
      ?estado=  id del Estado
    """
    qs = Dispositivo.objects.select_related(
        'g212_tipo', 'g212_marca', 'g212_propietario',
        'g212_estado', 'g212_co', 'g212_departamento', 'g212_municipio'
    ).exclude(g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS)

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(g212_serial__icontains=q) |
            Q(g212_propietario__g203_propietario__icontains=q) |
            Q(g212_marca__g202_marca__icontains=q)
        )

    tipo_id = request.GET.get('tipo')
    if tipo_id:
        qs = qs.filter(g212_tipo_id=tipo_id)

    estado_id = request.GET.get('estado')
    if estado_id:
        qs = qs.filter(g212_estado_id=estado_id)

    # Excluir dispositivos ya asignados a OTROS colaboradores.
    # Si se pasa ?colaborador_id=X se permite mostrar también los ya asignados
    # al propio colaborador X (para que el modal no los oculte).
    if request.GET.get('solo_disponibles'):
        colaborador_id = request.GET.get('colaborador_id')
        ya_asignados_qs = AsignacionColaborador.objects.values_list('g216_dispositivo_id', flat=True)
        if colaborador_id:
            # Excluir solo los asignados a OTROS colaboradores distintos a este
            ya_asignados_qs = ya_asignados_qs.exclude(g216_colaborador_id=colaborador_id)
        qs = qs.exclude(g212_id__in=ya_asignados_qs)

    dispositivos = []
    for d in qs.order_by('g212_serial'):
        dispositivos.append({
            'id':          d.g212_id,
            'serial':      d.g212_serial,
            'tipo':        d.g212_tipo.g200_tipo_dispositivo if d.g212_tipo else '—',
            'tipo_id':     d.g212_tipo_id,
            'marca':       d.g212_marca.g202_marca if d.g212_marca else '—',
            'marca_id':    d.g212_marca_id,
            'propietario': d.g212_propietario.g203_propietario if d.g212_propietario else '—',
            'propietario_id': d.g212_propietario_id,
            'estado':      d.g212_estado.g201_descripcion if d.g212_estado else '—',
            'estado_id':   d.g212_estado_id,
            'co':          f"{d.g212_co.g207_co} — {d.g212_co.g207_descripcion_co}" if d.g212_co else '—',
            'co_id':       d.g212_co_id,
            'nombre_equipo': d.g212_nombre_equipo or '—',
            'valor_promedio':      str(d.g212_valor_promedio) if d.g212_valor_promedio else None,
            'valor_arrendamiento': str(d.g212_valor_arrendamiento) if d.g212_valor_arrendamiento else None,
            'departamento':    d.g212_departamento.g204_departamento if d.g212_departamento else '—',
            'departamento_id': d.g212_departamento_id,
            'municipio':       d.g212_municipio.g205_municipio if d.g212_municipio else '—',
            'municipio_id':    d.g212_municipio_id,
            'observaciones':   d.g212_observaciones or '',
            'fecha_registro':  d.g212_fecha_registro.strftime('%d/%m/%Y %H:%M'),
        })

    # Estadísticas rápidas: total + conteo por CADA estado existente
    total = Dispositivo.objects.count()
    estados_conteo = list(
        Dispositivo.objects
        .values('g212_estado__g201_descripcion')
        .annotate(cantidad=Count('g212_id'))
        .order_by('g212_estado__g201_descripcion')
    )
    stats = {
        'total': total,
        'por_estado': [
            {
                'estado': row['g212_estado__g201_descripcion'] or 'SIN ESTADO',
                'cantidad': row['cantidad'],
            }
            for row in estados_conteo
        ],
    }

    return _json_ok({'dispositivos': dispositivos, 'stats': stats})

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_dispositivo_detalle(request, pk):
    """Detalle completo de un dispositivo incluyendo sus características."""
    d = get_object_or_404(
        Dispositivo.objects.select_related(
            'g212_tipo', 'g212_marca', 'g212_propietario',
            'g212_estado', 'g212_co', 'g212_departamento', 'g212_municipio'
        ),
        pk=pk
    )
    data = {
        'id':          d.g212_id,
        'serial':      d.g212_serial,
        'tipo':        d.g212_tipo.g200_tipo_dispositivo if d.g212_tipo else '—',
        'tipo_id':     d.g212_tipo_id,
        'marca':       d.g212_marca.g202_marca if d.g212_marca else '—',
        'marca_id':    d.g212_marca_id,

        'propietario': d.g212_propietario.g203_propietario if d.g212_propietario else '—',
        'propietario_id': d.g212_propietario_id,
        'estado':      d.g212_estado.g201_descripcion if d.g212_estado else '—',
        'estado_id':   d.g212_estado_id,
        'co':          f"{d.g212_co.g207_co} — {d.g212_co.g207_descripcion_co}" if d.g212_co else '—',
        'co_id':       d.g212_co_id,
        'nombre_equipo': d.g212_nombre_equipo or '',
        'valor_promedio':      str(d.g212_valor_promedio) if d.g212_valor_promedio else '',
        'valor_arrendamiento': str(d.g212_valor_arrendamiento) if d.g212_valor_arrendamiento else '',
        'departamento_id': d.g212_departamento_id,
        'municipio_id':    d.g212_municipio_id,
        'observaciones':   d.g212_observaciones or '',
        'departamento':    d.g212_departamento.g204_departamento if d.g212_departamento else '—',
        'municipio':       d.g212_municipio.g205_municipio if d.g212_municipio else '—',
        'caracteristicas': _get_caracteristicas(d),
        'asignado_a':  AsignacionColaborador.objects.filter(g216_dispositivo=d).select_related('g216_colaborador').first().__class__ and (lambda a: a.g216_colaborador.g215_nombre if a else None)(AsignacionColaborador.objects.filter(g216_dispositivo=d).select_related('g216_colaborador').first()),
    }
    return _json_ok(data)


def _get_caracteristicas(d):
    """Retorna las características específicas según el tipo de dispositivo."""
    tipo = d.g212_tipo.g200_tipo_dispositivo if d.g212_tipo else ''

    if tipo in ('TORRE DE ESCRITORIO', 'PORTATIL'):
        try:
            pc = d.caract_pc
            return {
                'grupo': 'pc',
                'procesador_id':    pc.g222_procesador_id,
                'procesador':       pc.g222_procesador.g209_procesador if pc.g222_procesador else '',
                'so_id':            pc.g222_so_id,
                'so':               pc.g222_so.g210_so if pc.g222_so else '',
                'antivirus_id':     pc.g222_antivirus_id,
                'antivirus':        pc.g222_antivirus.g208_antivirus if pc.g222_antivirus else '',
                'licencia_id':      pc.g222_licencia_id,
                'licencia':         pc.g222_licencia.g211_office if pc.g222_licencia else '',
                'correo_office':    pc.g222_correo_office or '',
                'key_office':       pc.g222_key_office or '',
                'ram':              pc.g222_ram,
                'tipo_disco_id':    pc.g222_tipo_disco_id,
                'tipo_disco':       pc.g222_tipo_disco.g231_tipo_disco if pc.g222_tipo_disco else '',
                'almacenamiento_id': pc.g222_almacenamiento_id,
                'almacenamiento':   pc.g222_almacenamiento.g219_almacenamiento if pc.g222_almacenamiento else '',
                'activo':           pc.g222_activo or '',
                'pulgadas':         str(pc.g222_pulgadas) if pc.g222_pulgadas else '',
                'nombre_equipo':      d.g212_nombre_equipo or '',
                'valor_promedio':     str(d.g212_valor_promedio) if d.g212_valor_promedio else '',
                'valor_arrendamiento': str(d.g212_valor_arrendamiento) if d.g212_valor_arrendamiento else '',
            }
        except Exception:
            return {'grupo': 'pc'}

    elif tipo in ('CELULAR', 'TABLET', 'MODEM WIFI', 'SIMCARD', 'TELEFONO FIJO'):
        try:
            mov = d.caract_movil
            return {
                'grupo':            'movil',
                'numero_linea':     mov.g223_numero_linea or '',
                'operador_id':      mov.g223_operador_id,
                'operador':         mov.g223_operador.g221_operador if mov.g223_operador else '',
                'plan_datos':       mov.g223_plan_datos or '',
                'imei1':            mov.g223_imei1 or '',
                'imei2':            mov.g223_imei2 or '',
                'cuenta_gmail':     mov.g223_cuenta_gmail or '',
                'contrasena_gmail': mov.g223_contrasena_gmail or '',
                'pulgadas':         str(mov.g223_pulgadas) if mov.g223_pulgadas else '',
                'almacenamiento_id':  mov.g223_almacenamiento_id,
                'almacenamiento':     mov.g223_almacenamiento.g219_almacenamiento if mov.g223_almacenamiento else '',
                'valor_promedio':     str(d.g212_valor_promedio) if d.g212_valor_promedio else '',
                'valor_arrendamiento': str(d.g212_valor_arrendamiento) if d.g212_valor_arrendamiento else '',
            }
        except Exception:
            return {'grupo': 'movil'}

    elif tipo == 'PANTALLA':
        try:
            pan = d.caract_pantalla
            return {
                'grupo':     'pantalla',
                'pulgadas':  str(pan.g224_pulgadas) if pan.g224_pulgadas else '',
                'resolucion': pan.g224_resolucion or '',
                'valor_promedio':     str(d.g212_valor_promedio) if d.g212_valor_promedio else '',
                'valor_arrendamiento': str(d.g212_valor_arrendamiento) if d.g212_valor_arrendamiento else '',
            }
        except Exception:
            return {'grupo': 'pantalla'}

    elif tipo == 'IMPRESORA':
        try:
            imp = d.caract_impresora
            return {
                'grupo':            'impresora',
                'tipo_impresora_id': imp.g225_tipo_impresora_id,
                'tipo_impresora':   imp.g225_tipo_impresora.g229_tipo_impresora if imp.g225_tipo_impresora else '',
                'funcion':          imp.g225_funcion or '',
            }
        except Exception:
            return {'grupo': 'impresora'}

    elif tipo == 'PERIFERICO':
        try:
            per = d.caract_periferico
            return {
                'grupo':              'periferico',
                'incluye_base':       per.g226_incluye_base,
                'incluye_teclado':    per.g226_incluye_teclado,
                'incluye_mouse':      per.g226_incluye_mouse,
                'incluye_auriculares': per.g226_incluye_auriculares,
                'incluye_cargador':   per.g226_incluye_cargador,
                'descripcion_adicional': per.g226_descripcion_adicional or '',
            }
        except Exception:
            return {'grupo': 'periferico'}

    elif tipo == 'LICENCIA OFFICE':
        try:
            lic = d.caract_licencia
            return {
                'grupo':    'licencia',
                'software': lic.g227_software or '',
                'version':  lic.g227_version or '',
                'key':      lic.g227_key or '',
                'correo':   lic.g227_correo or '',
                'fecha_vencimiento': str(lic.g227_fecha_vencimiento) if lic.g227_fecha_vencimiento else '',
                'almacenamiento_id': lic.g227_almacenamiento_id,
                'almacenamiento':    lic.g227_almacenamiento.g219_almacenamiento if lic.g227_almacenamiento else '',
                'valor_arrendamiento': str(d.g212_valor_arrendamiento) if d.g212_valor_arrendamiento else '',
            }
        except Exception:
            return {'grupo': 'licencia'}
        
        
    elif tipo== 'VIDEO BEAM':
        try:
            vb= d.caract_videobeam
            return{
                'grupo':            'videobeam',
                'lumenes':          str(vb.g232_lumenes) if vb.g232_lumenes else '',
                
                
                }
        except Exception:
            return {'grupo': 'videobeam'}
    return {}


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_verificar_serial(request):
    serial = request.GET.get('serial', '').strip()
    if not serial:
        return _json_err('Serial requerido', 400)
    d = Dispositivo.objects.filter(g212_serial=serial).first()
    if not d:
        return _json_ok({'existe': False, 'asignado_a': None})
    asig = AsignacionColaborador.objects.filter(
        g216_dispositivo=d
    ).select_related('g216_colaborador').first()
    return _json_ok({
        'existe':         True,
        'asignado_a':     asig.g216_colaborador.g215_nombre if asig else None,
        'colaborador_id': asig.g216_colaborador.g215_id if asig else None,
    })

SERIAL_PREFIJO_REGEX_CACHE = {}


def _regex_para_prefijo(prefijo):
    """Cachea el regex compilado por prefijo para no recompilar en cada llamada."""
    if prefijo not in SERIAL_PREFIJO_REGEX_CACHE:
        SERIAL_PREFIJO_REGEX_CACHE[prefijo] = re.compile(
            rf'^{re.escape(prefijo)}(\d+)$', re.IGNORECASE
        )
    return SERIAL_PREFIJO_REGEX_CACHE[prefijo]


# Tipos de dispositivo con serial autogenerado: nombre EXACTO del tipo
# (tal como está en g200_tipo_dispositivo) -> (prefijo, cantidad de dígitos).
# padding=0 -> sin ceros a la izquierda (L186). padding=3 -> P001, P002...
SERIES_AUTOMATICAS = {
    'LICENCIA OFFICE': ('L', 0),
    'PERIFERICO':       ('P', 3),
}


def _calcular_siguiente_num(prefijo, seriales):
    """Dado un iterable de seriales y un prefijo, devuelve el próximo número."""
    regex = _regex_para_prefijo(prefijo)
    max_num = 0
    for s in seriales:
        m = regex.match(s.strip())
        if m:
            max_num = max(max_num, int(m.group(1)))
    return max_num + 1


def _formatear_serial(prefijo, numero, padding):
    numero_str = str(numero).zfill(padding) if padding else str(numero)
    return f'{prefijo}{numero_str}'


def _normalizar_serial(valor):
    """Quita todos los espacios en blanco (incluso internos) del serial y lo
    pasa a mayúsculas, para que 'ML- 167904', 'ml-167904' y 'ML-167904' se
    traten como el mismo valor."""
    return re.sub(r'\s+', '', valor or '').upper()


def _generar_siguiente_serial(prefijo, padding):
    """
    Calcula el siguiente serial secuencial para el prefijo dado
    (ej. LICENCIA OFFICE -> 'L', 0  =>  L185 -> L186)
    (ej. PERIFERICO      -> 'P', 3  =>  (nada) -> P001, P001 -> P002)

    IMPORTANTE: debe llamarse dentro de un transaction.atomic() activo,
    ya que bloquea (select_for_update) las filas con ese prefijo para
    evitar que dos creaciones simultáneas generen el mismo serial.
    """
    seriales = (
        Dispositivo.objects
        .select_for_update()
        .filter(g212_serial__istartswith=prefijo)
        .values_list('g212_serial', flat=True)
    )
    siguiente = _calcular_siguiente_num(prefijo, seriales)
    return _formatear_serial(prefijo, siguiente, padding)


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_siguiente_serial(request):
    """
    Devuelve, SOLO como vista previa (sin bloquear filas), cuál sería el
    próximo serial automático para el tipo de dispositivo indicado
    (?tipo_id=<id>), para mostrarlo en el formulario mientras el usuario
    llena los demás datos.

    Si el tipo no tiene serie automática configurada (SERIES_AUTOMATICAS),
    responde {'aplica': false} y el frontend deja el campo editable normal.

    El valor DEFINITIVO se recalcula con lock justo al crear el
    dispositivo (api_dispositivo_crear), así que si hay una carrera entre
    dos usuarios el serial final puede diferir del que se mostró aquí —
    por eso el campo en el formulario queda de solo lectura pero el valor
    real se re-verifica en el servidor al guardar.
    """
    tipo_id = request.GET.get('tipo_id')
    tipo_obj = TipoDispositivo.objects.filter(pk=tipo_id).first() if tipo_id else None
    tipo_nombre = tipo_obj.g200_tipo_dispositivo.strip().upper() if tipo_obj else ''

    config = SERIES_AUTOMATICAS.get(tipo_nombre)
    if not config:
        return _json_ok({'aplica': False})

    prefijo, padding = config
    seriales = Dispositivo.objects.filter(
        g212_serial__istartswith=prefijo
    ).values_list('g212_serial', flat=True)

    siguiente = _calcular_siguiente_num(prefijo, seriales)
    serial = _formatear_serial(prefijo, siguiente, padding)
    return _json_ok({'aplica': True, 'serial': serial})


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_dispositivo_crear(request):
    """Crea un nuevo dispositivo con sus características."""
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')

    required = ['tipo_id', 'propietario_id', 'co_id', 'estado_id',
                'departamento_id', 'municipio_id']
    for field in required:
        if not body.get(field):
            return _json_err(f'Campo requerido: {field}')

    try:
        with transaction.atomic():
            tipo_obj = TipoDispositivo.objects.filter(pk=body['tipo_id']).first()
            tipo_nombre = tipo_obj.g200_tipo_dispositivo.strip().upper() if tipo_obj else ''
            config_serie = SERIES_AUTOMATICAS.get(tipo_nombre)

            # Generar serial automático si no viene
            serial = _normalizar_serial(body.get('serial', ''))
            if not serial:
                if config_serie:
                    prefijo, padding = config_serie
                    serial = _generar_siguiente_serial(prefijo, padding)
                else:
                    ultimo = Dispositivo.objects.order_by('-g212_id').first()
                    siguiente_id = (ultimo.g212_id + 1) if ultimo else 1
                    serial = str(siguiente_id).zfill(5)
            elif Dispositivo.objects.filter(g212_serial__iexact=serial).exists():
                return _json_err(f'Ya existe un dispositivo con el serial "{serial}"')

            d = Dispositivo.objects.create(
                g212_serial=serial,
                g212_tipo_id=body['tipo_id'],
                g212_marca_id=body.get('marca_id') or None,
                g212_propietario_id=body['propietario_id'],
                g212_estado_id=body['estado_id'],
                g212_co_id=body['co_id'],
                g212_nombre_equipo=(body.get('nombre_equipo') or '').strip().upper(),
                g212_valor_promedio=body.get('valor_promedio') or None,
                g212_valor_arrendamiento=body.get('valor_arrendamiento') or None,
                g212_departamento_id=body['departamento_id'],
                g212_municipio_id=body['municipio_id'],
                g212_observaciones=body.get('observaciones', ''),
            )
            _save_caracteristicas(d, body)

    except Exception as e:
        return _json_err(str(e))

    return _json_ok({'id': d.g212_id, 'serial': d.g212_serial})

@login_required(login_url='login')
@require_http_methods(['PUT'])
def api_dispositivo_editar(request, pk):
    """Edita un dispositivo existente."""
    d = get_object_or_404(Dispositivo, pk=pk)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')

    try:
        with transaction.atomic():
            estado_anterior_desc = d.g212_estado.g201_descripcion if d.g212_estado else None
            nuevo_serial = _normalizar_serial(body.get('serial', d.g212_serial))
            if (nuevo_serial and nuevo_serial.lower() != d.g212_serial.lower()
                    and Dispositivo.objects.filter(g212_serial__iexact=nuevo_serial).exclude(pk=d.pk).exists()):
                return _json_err(f'Ya existe un dispositivo con el serial "{nuevo_serial}"')
            d.g212_serial = nuevo_serial or d.g212_serial
            d.g212_tipo_id = body.get('tipo_id', d.g212_tipo_id)
            d.g212_marca_id = body.get('marca_id') or None
            d.g212_propietario_id = body.get('propietario_id', d.g212_propietario_id)
            d.g212_estado_id = body.get('estado_id', d.g212_estado_id)
            d.g212_co_id = body.get('co_id') or None
            # Nombre/valor solo se actualizan si llega un valor real: estos campos
            # se envían desde la sección "Características" (según tipo), así que
            # un cuerpo sin ese dato NO debe borrar lo que ya estaba guardado
            # (ej. cargado por Excel) en dispositivos sin esa sección.
            if body.get('nombre_equipo'):
                d.g212_nombre_equipo = body['nombre_equipo'].strip().upper()
            if body.get('valor_promedio') not in (None, ''):
                d.g212_valor_promedio = body['valor_promedio']
            if body.get('valor_arrendamiento') not in (None, ''):
                d.g212_valor_arrendamiento = body['valor_arrendamiento']
            d.g212_departamento_id = body.get('departamento_id') or None
            d.g212_municipio_id = body.get('municipio_id') or None
            d.g212_observaciones = body.get('observaciones', d.g212_observaciones)
            d.save()
            _save_caracteristicas(d, body)
            responsable = _nombre_completo_usuario(request)
            _registrar_historial_auto(
                dispositivo    = d,
                nombre_novedad = 'EDICIÓN',
                responsable    = responsable,
                observaciones  = 'Dispositivo editado desde Inventario.',
                co             = d.g212_co,
            )
            _registrar_transicion_inactivo(d, estado_anterior_desc, responsable, co=d.g212_co)

            estado_nuevo_desc = d.g212_estado.g201_descripcion if d.g212_estado else None
            if estado_nuevo_desc in ESTADOS_INACTIVOS and body.get('checklist'):
                _guardar_checklist_dispositivo(d, body['checklist'], responsable)
    except Exception as e:
        return _json_err(str(e))

    return _json_ok({'id': d.g212_id, 'serial': d.g212_serial})


def _save_caracteristicas(d, body):
    """
    Guarda o actualiza las características según el grupo que venga en body['caract'].
    body['caract'] = {'grupo': 'pc'|'movil'|'pantalla'|..., ...campos}
    """
    caract = body.get('caract', {})
    grupo = caract.get('grupo')

    if grupo == 'pc':
        CaracteristicaPC.objects.update_or_create(
            g222_dispositivo=d,
            defaults={
                'g222_procesador_id':    caract.get('procesador_id') or None,
                'g222_so_id':            caract.get('so_id') or None,
                'g222_antivirus_id':     caract.get('antivirus_id') or None,
                'g222_licencia_id':      caract.get('licencia_id') or None,
                'g222_correo_office':    caract.get('correo_office', ''),
                'g222_key_office':       caract.get('key_office', ''),
                'g222_ram':              _normalizar_ram(caract.get('ram')),
                'g222_tipo_disco_id':    caract.get('tipo_disco_id') or None,
                'g222_almacenamiento_id': caract.get('almacenamiento_id') or None,
                'g222_activo':           caract.get('activo', ''),
                'g222_pulgadas':         caract.get('pulgadas') or None,
            }
        )
        # Guardar nombre_equipo, valor_promedio y valor_arrendamiento (campos de Dispositivo)
        # que vienen dentro de la sección Características de TORRE/PORTÁTIL.
        # Se usa 'in caract' (no 'is not None') para vp/va: así, si el usuario
        # deja el campo en blanco a propósito, se guarda vacío en vez de
        # conservar para siempre el valor anterior.
        nombre = caract.get('nombre_equipo')
        tiene_vp = 'valor_promedio' in caract
        tiene_va = 'valor_arrendamiento' in caract
        if nombre or tiene_vp or tiene_va:
            if nombre:
                d.g212_nombre_equipo = nombre.strip().upper()
            if tiene_vp:
                d.g212_valor_promedio = caract.get('valor_promedio') or None
            if tiene_va:
                d.g212_valor_arrendamiento = caract.get('valor_arrendamiento') or None
            d.save()

    elif grupo == 'movil':
        CaracteristicaMovil.objects.update_or_create(
            g223_dispositivo=d,
            defaults={
                'g223_numero_linea':   caract.get('numero_linea', ''),
                'g223_operador_id':    caract.get('operador_id') or None,
                'g223_plan_datos':     caract.get('plan_datos', ''),
                'g223_imei1':          caract.get('imei1', ''),
                'g223_imei2':          caract.get('imei2', ''),
                'g223_cuenta_gmail':   caract.get('cuenta_gmail', ''),
                'g223_contrasena_gmail': caract.get('contrasena_gmail', ''),
                'g223_pulgadas':       caract.get('pulgadas') or None,
                'g223_almacenamiento_id': caract.get('almacenamiento_id') or None,
            }
        )
        # Guardar valor_promedio y valor_arrendamiento si vienen en caract (MODEM, SIMCARD, TABLET)
        tiene_vp = 'valor_promedio' in caract
        tiene_va = 'valor_arrendamiento' in caract
        if tiene_vp or tiene_va:
            if tiene_vp:
                d.g212_valor_promedio = caract.get('valor_promedio') or None
            if tiene_va:
                d.g212_valor_arrendamiento = caract.get('valor_arrendamiento') or None
            d.save()

    elif grupo == 'pantalla':
        CaracteristicaPantalla.objects.update_or_create(
            g224_dispositivo=d,
            defaults={
                'g224_pulgadas':   caract.get('pulgadas') or None,
                'g224_resolucion': caract.get('resolucion', ''),
            }
        )
        # Guardar valor_promedio y valor_arrendamiento si vienen en caract (PANTALLA)
        tiene_vp = 'valor_promedio' in caract
        tiene_va = 'valor_arrendamiento' in caract
        if tiene_vp or tiene_va:
            if tiene_vp:
                d.g212_valor_promedio = caract.get('valor_promedio') or None
            if tiene_va:
                d.g212_valor_arrendamiento = caract.get('valor_arrendamiento') or None
            d.save()

    elif grupo == 'impresora':
        # 'funcion' no tiene campo en el formulario manual (solo llega por
        # carga masiva) — si no viene en el body, se preserva lo existente
        # en vez de sobreescribir con vacío.
        existente = CaracteristicaImpresora.objects.filter(g225_dispositivo=d).first()
        funcion = caract.get('funcion') or (existente.g225_funcion if existente else '')
        CaracteristicaImpresora.objects.update_or_create(
            g225_dispositivo=d,
            defaults={
                'g225_tipo_impresora_id': caract.get('tipo_impresora_id') or None,
                'g225_funcion':           funcion,
            }
        )

    elif grupo == 'periferico':
        # 'descripcion_adicional' no tiene campo en el formulario manual
        # (solo llega por carga masiva) — se preserva si no viene en el body.
        existente = CaracteristicaPeriferico.objects.filter(g226_dispositivo=d).first()
        descripcion = caract.get('descripcion_adicional') or (existente.g226_descripcion_adicional if existente else '')
        CaracteristicaPeriferico.objects.update_or_create(
            g226_dispositivo=d,
            defaults={
                'g226_incluye_base':         caract.get('incluye_base', False),
                'g226_incluye_teclado':      caract.get('incluye_teclado', False),
                'g226_incluye_mouse':        caract.get('incluye_mouse', False),
                'g226_incluye_auriculares':  caract.get('incluye_auriculares', False),
                'g226_incluye_cargador':     caract.get('incluye_cargador', False),
                'g226_descripcion_adicional': descripcion,
            }
        )

    elif grupo == 'licencia':
        CaracteristicaLicencia.objects.update_or_create(
            g227_dispositivo=d,
            defaults={
                'g227_software': caract.get('software', ''),
                'g227_version':  caract.get('version', ''),
                'g227_key':      caract.get('key', ''),
                'g227_correo':   caract.get('correo', ''),
                'g227_fecha_vencimiento': caract.get('fecha_vencimiento') or None,
                'g227_almacenamiento_id': caract.get('almacenamiento_id') or None,
            }
        )
        if 'valor_arrendamiento' in caract:
            d.g212_valor_arrendamiento = caract.get('valor_arrendamiento') or None
            d.save()

    elif grupo== 'videobeam':
        
        
        CaracteristicasVideoBeam.objects.update_or_create(
            g232_dispositivo=d,
            defaults={
                
                 'g232_lumenes':             caract.get('lumenes') or None,
                
            }
            
           
        )
        
        

@login_required(login_url='login')
@require_http_methods(['DELETE'])
def api_dispositivo_eliminar(request, pk):
    """Elimina un dispositivo permanentemente."""
    d = get_object_or_404(Dispositivo, pk=pk)
    serial = d.g212_serial
    _registrar_historial_auto(
        dispositivo    = d,
        nombre_novedad = 'ELIMINACIÓN PERMANENTE',
        responsable    = request.user.get_full_name() or request.user.username,
        observaciones  = f'Dispositivo eliminado permanentemente desde Inventario.',
        co             = d.g212_co,
    )
    d.delete()
    return _json_ok({'serial': serial})


# ═══════════════════════════════════════════════════
#  HISTORIAL DE EQUIPOS
# ═══════════════════════════════════════════════════
@login_required(login_url='login')
@require_http_methods(['GET'])
def api_historial(request):
    """
    Retorna historial filtrable por tipo y/o serial.
      ?tipo_id=  id del TipoDispositivo
      ?serial=   texto parcial del serial
    """
    qs = HistorialEquipo.objects.select_related(
    'g214_dispositivo__g212_tipo',
    'g214_dispositivo__g212_marca',
    'g214_dispositivo__g212_propietario',
    'g214_dispositivo__g212_estado',
    'g214_dispositivo__g212_co',
    'g214_novedad',
    'g214_co',
    ).order_by('-g214_fecha', '-g214_hora')

    tipo_id = request.GET.get('tipo_id')
    if tipo_id:
        qs = qs.filter(g214_dispositivo__g212_tipo_id=tipo_id)

    serial = request.GET.get('serial', '').strip()
    if serial:
        # Incluye también dispositivos ya eliminados: para esos no queda
        # relación en vivo, así que se busca en la instantánea de texto
        # (g214_dispositivo_desc) guardada en el momento del evento.
        qs = qs.filter(
            Q(g214_dispositivo__g212_serial__icontains=serial) |
            Q(g214_dispositivo__isnull=True, g214_dispositivo_desc__icontains=serial)
        )

    registros = []
    for h in qs:
        dispositivo = h.g214_dispositivo
        if dispositivo:
            registros.append({
                'id':          h.g214_id,
                'serial':      dispositivo.g212_serial,
                'tipo':        dispositivo.g212_tipo.g200_tipo_dispositivo if dispositivo.g212_tipo else '—',
                'marca':       dispositivo.g212_marca.g202_marca if dispositivo.g212_marca else '—',
                'propietario': dispositivo.g212_propietario.g203_propietario if dispositivo.g212_propietario else '—',
                'estado':      dispositivo.g212_estado.g201_descripcion if dispositivo.g212_estado else '—',
                'co_equipo':   f"{dispositivo.g212_co.g207_co}" if dispositivo.g212_co else '—',
                'novedad':     h.g214_novedad.g220_novedad if h.g214_novedad else '—',
                'novedad_id':  h.g214_novedad_id,
                'fecha':       h.g214_fecha.strftime('%Y-%m-%d'),
                'hora':        h.g214_hora.strftime('%H:%M'),
                'responsable': h.g214_responsable,
                'co':          f"{h.g214_co.g207_co} — {h.g214_co.g207_descripcion_co}" if h.g214_co else '—',
                'observaciones': h.g214_observaciones or '',
            })
        else:
            registros.append({
                'id':          h.g214_id,
                'serial':      h.g214_dispositivo_desc or '(dispositivo eliminado)',
                'tipo':        '—',
                'marca':       '—',
                'propietario': '—',
                'estado':      'ELIMINADO',
                'co_equipo':   '—',
                'novedad':     h.g214_novedad.g220_novedad if h.g214_novedad else '—',
                'novedad_id':  h.g214_novedad_id,
                'fecha':       h.g214_fecha.strftime('%Y-%m-%d'),
                'hora':        h.g214_hora.strftime('%H:%M'),
                'responsable': h.g214_responsable,
                'co':          f"{h.g214_co.g207_co} — {h.g214_co.g207_descripcion_co}" if h.g214_co else '—',
                'observaciones': h.g214_observaciones or '',
            })

    return _json_ok(registros)



def _nombre_completo_usuario(request):
    """
    Nombre real del usuario logueado, para dejarlo como 'Registrado por' en vez
    de su cédula. El login usa la cédula como username, así que se busca el
    Colaborador correspondiente; si no existe, se usa get_full_name()/username
    como respaldo (igual que antes).
    """
    colaborador = Colaborador.objects.filter(g215_documento=request.user.username).first()
    if colaborador:
        return colaborador.g215_nombre
    return request.user.get_full_name() or request.user.username


#  Registrar historial automático

def _registrar_historial_auto(dispositivo, nombre_novedad, responsable, observaciones='', co=None):
    """
    Crea un registro en HistorialEquipo automáticamente.
    Crea el TipoNovedad si no existe.

    Siempre guarda una instantánea de texto (serial — tipo — marca) del
    dispositivo en g214_dispositivo_desc, para que el registro se pueda
    seguir identificando aunque el dispositivo sea eliminado permanentemente
    más adelante (g214_dispositivo pasaría a NULL por on_delete=SET_NULL).
    """
    from datetime import date, datetime
    novedad_obj, _ = TipoNovedad.objects.get_or_create(
        g220_novedad__iexact=nombre_novedad,
        defaults={'g220_novedad': nombre_novedad, 'g220_estado': True}
    )
    tipo_nombre = dispositivo.g212_tipo.g200_tipo_dispositivo if dispositivo.g212_tipo else '—'
    marca_nombre = dispositivo.g212_marca.g202_marca if dispositivo.g212_marca else '—'
    HistorialEquipo.objects.create(
        g214_dispositivo      = dispositivo,
        g214_dispositivo_desc = f"{dispositivo.g212_serial} — {tipo_nombre} — {marca_nombre}",
        g214_novedad       = novedad_obj,
        g214_fecha         = date.today(),
        g214_hora          = datetime.now().time(),
        g214_responsable   = responsable or 'Sistema',
        g214_co            = co,
        g214_observaciones = observaciones,
    )


def _registrar_transicion_inactivo(dispositivo, estado_anterior_desc, responsable, co=None):
    """
    Si el dispositivo entró o salió del grupo de estados inactivos
    (ESTADOS_INACTIVOS), registra el movimiento en Historial de Equipo.
    Se llama después de guardar el nuevo estado en `dispositivo`.
    """
    estado_nuevo_desc = dispositivo.g212_estado.g201_descripcion if dispositivo.g212_estado else None
    era_inactivo = estado_anterior_desc in ESTADOS_INACTIVOS
    es_inactivo  = estado_nuevo_desc in ESTADOS_INACTIVOS
    if era_inactivo == es_inactivo:
        return
    if es_inactivo:
        _registrar_historial_auto(
            dispositivo    = dispositivo,
            nombre_novedad = f'PASÓ A INACTIVO ({estado_nuevo_desc})',
            responsable    = responsable,
            observaciones  = f'El dispositivo cambió de estado "{estado_anterior_desc or "—"}" a "{estado_nuevo_desc}" y salió del Inventario activo.',
            co             = co,
        )
    else:
        _registrar_historial_auto(
            dispositivo    = dispositivo,
            nombre_novedad = 'REACTIVADO',
            responsable    = responsable,
            observaciones  = f'El dispositivo cambió de estado "{estado_anterior_desc}" a "{estado_nuevo_desc or "—"}" y volvió al Inventario activo.',
            co             = co,
        )


def _guardar_checklist_dispositivo(dispositivo, checklist_body, responsable):
    """
    Guarda el Checklist de Inventario de un dispositivo (cabecera + respuestas)
    y deja constancia en Historial de Equipo. Se llama desde api_dispositivo_editar
    cuando el dispositivo entra a un estado inactivo (ver ESTADOS_INACTIVOS).
    """
    tipo_nombre = dispositivo.g212_tipo.g200_tipo_dispositivo if dispositivo.g212_tipo else '—'
    marca_nombre = dispositivo.g212_marca.g202_marca if dispositivo.g212_marca else '—'
    cd = ChecklistDispositivo.objects.create(
        g237_dispositivo      = dispositivo,
        g237_dispositivo_desc = f"{dispositivo.g212_serial} — {tipo_nombre} — {marca_nombre}",
        g237_responsable      = responsable,
        g237_observaciones    = checklist_body.get('observaciones', ''),
        g237_resp_nombre      = checklist_body.get('resp_nombre', ''),
        g237_resp_cedula      = checklist_body.get('resp_cedula', ''),
        g237_resp_area        = checklist_body.get('resp_area', ''),
        g237_resp_cargo       = checklist_body.get('resp_cargo', ''),
    )
    for r in checklist_body.get('respuestas', []):
        item = ItemChecklist.objects.filter(pk=r.get('item_id')).first()
        RespuestaChecklist.objects.create(
            g238_checklist   = cd,
            g238_item        = item,
            g238_item_desc   = item.g236_pregunta if item else (r.get('pregunta') or ''),
            g238_respuesta   = bool(r.get('respuesta')),
            g238_valor_texto = r.get('valor_texto') or None,
        )
    _registrar_historial_auto(
        dispositivo    = dispositivo,
        nombre_novedad = 'CHECKLIST REALIZADO',
        responsable    = responsable,
        observaciones  = 'Se completó el checklist de inventario al cambiar el estado del equipo.',
        co             = dispositivo.g212_co,
    )


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_historial_por_dispositivo(request, dispositivo_id):
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    registros = HistorialEquipo.objects.select_related(
        'g214_novedad', 'g214_co'
    ).filter(g214_dispositivo=dispositivo).order_by('-g214_fecha', '-g214_hora')

    historial = []
    for h in registros:
        historial.append({
            'id':            h.g214_id,
            'novedad':       h.g214_novedad.g220_novedad if h.g214_novedad else '—',
            'fecha':         h.g214_fecha.strftime('%d/%m/%Y'),
            'hora':          h.g214_hora.strftime('%H:%M'),
            'responsable':   h.g214_responsable,
            'co':            f"{h.g214_co.g207_co} — {h.g214_co.g207_descripcion_co}" if h.g214_co else '—',
            'observaciones': h.g214_observaciones or '',
            'fecha_registro': h.g214_fecha_registro.strftime('%d/%m/%Y %H:%M'),
        })

    return _json_ok({
        'dispositivo': {
            'id':          dispositivo.g212_id,
            'serial':      dispositivo.g212_serial,
            'tipo':        dispositivo.g212_tipo.g200_tipo_dispositivo if dispositivo.g212_tipo else '—',
            'marca':       dispositivo.g212_marca.g202_marca if dispositivo.g212_marca else '—',
            'propietario': dispositivo.g212_propietario.g203_propietario if dispositivo.g212_propietario else '—',
            'estado':      dispositivo.g212_estado.g201_descripcion if dispositivo.g212_estado else '—',
            'co':          dispositivo.g212_co.g207_co if dispositivo.g212_co else '—',
        },
        'historial': historial,
        'total':     len(historial),
    })


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_historial_crear(request):
    """Crea un nuevo registro de historial."""
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')

    required = ['dispositivo_id', 'novedad_id', 'fecha', 'hora', 'responsable']
    for f in required:
        if not body.get(f):
            return _json_err(f'Campo requerido: {f}')

    try:
        h = HistorialEquipo.objects.create(
            g214_dispositivo_id=body['dispositivo_id'],
            g214_novedad_id=body['novedad_id'],
            g214_fecha=body['fecha'],
            g214_hora=body['hora'],
            g214_responsable=body['responsable'],
            g214_co_id=body.get('co_id') or None,
            g214_observaciones=body.get('observaciones', ''),
        )
    except Exception as e:
        return _json_err(str(e))

    return _json_ok({'id': h.g214_id})



#  CENTRO DE COSTOS — Estadísticas agrupadas

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_centro_operaciones(request):
    """
    Consulta de dispositivos agrupados por tipo para Centro de Costos.
    El área real está en el COLABORADOR (g215_Area_id), no en el dispositivo.
    Flujo: Área -> Colaboradores de esa área -> Dispositivos asignados (j216)
      ?co_id=     id del Area / CentroCosto (j228_area)
      ?prop_id=   id del Propietario
      ?tipo_id=   id del TipoDispositivo
    """
    co_id   = request.GET.get('co_id')
    prop_id = request.GET.get('prop_id')
    tipo_id = request.GET.get('tipo_id')

    asign_qs = AsignacionColaborador.objects.select_related('g216_colaborador', 'g216_dispositivo')
    if co_id:
        asign_qs = asign_qs.filter(g216_colaborador__g215_Area_id=co_id)
    if prop_id:
        asign_qs = asign_qs.filter(g216_dispositivo__g212_propietario_id=prop_id)
    if tipo_id:
        asign_qs = asign_qs.filter(g216_dispositivo__g212_tipo_id=tipo_id)

    dispositivo_ids = asign_qs.values_list('g216_dispositivo_id', flat=True)

    qs = Dispositivo.objects.filter(g212_id__in=dispositivo_ids).select_related(
        'g212_tipo', 'g212_estado', 'g212_propietario'
    )

    from django.db.models import Sum
    from decimal import Decimal

    # Tipos que aplican BitDefender (solo PORTÁTIL y TORRE DE ESCRITORIO)
    TIPOS_CON_BITDEFENDER = {'PORTATIL', 'PORTÁTIL', 'TORRE DE ESCRITORIO'}
    COSTO_BITDEFENDER_POR_EQUIPO = Decimal('6000')

    # Totales generales
    total       = qs.count()
    habilitados = qs.filter(g212_estado__g201_descripcion='HABILITADO').count()
    otros       = total - habilitados

    # Agrupar por tipo — cantidad + suma de costos
    grupos_raw = (
        qs.values('g212_tipo__g200_tipo_dispositivo', 'g212_estado__g201_descripcion')
        .annotate(
            cantidad=Count('g212_id'),
            suma_arrendamiento=Sum('g212_valor_arrendamiento'),
            suma_promedio=Sum('g212_valor_promedio'),
        )
    )

    grupos = {}
    for row in grupos_raw:
        tipo_nombre   = row['g212_tipo__g200_tipo_dispositivo'] or 'SIN TIPO'
        estado_nombre = row['g212_estado__g201_descripcion'] or 'SIN ESTADO'
        cant          = row['cantidad']
        arr           = row['suma_arrendamiento'] or Decimal('0')
        prom          = row['suma_promedio'] or Decimal('0')

        if tipo_nombre not in grupos:
            grupos[tipo_nombre] = {
                'tipo': tipo_nombre,
                'cantidad': 0,
                'habilitados': 0, 'inhabilitados': 0, 'asignados': 0,
                'costo_mensual': Decimal('0'),
                'costo_promedio': Decimal('0'),
                'aplica_bitdefender': tipo_nombre.upper() in TIPOS_CON_BITDEFENDER,
            }
        grupos[tipo_nombre]['cantidad']       += cant
        grupos[tipo_nombre]['costo_mensual']  += arr
        grupos[tipo_nombre]['costo_promedio'] += prom
        if estado_nombre == 'HABILITADO':
            grupos[tipo_nombre]['habilitados'] += cant
        elif estado_nombre == 'INHABILITADO':
            grupos[tipo_nombre]['inhabilitados'] += cant
        elif estado_nombre == 'ASIGNADO':
            grupos[tipo_nombre]['asignados'] += cant

    # Calcular totales
    total_arrendamiento      = sum(g['costo_mensual'] for g in grupos.values())
    total_promedio           = sum(g['costo_promedio'] for g in grupos.values())
    total_bitdefender_global = Decimal('0')

    grupos_list = []
    for g in sorted(grupos.values(), key=lambda x: x['tipo']):
        bd_grupo = COSTO_BITDEFENDER_POR_EQUIPO * g['cantidad'] if g['aplica_bitdefender'] else Decimal('0')
        total_bitdefender_global += bd_grupo
        grupos_list.append({
            'tipo':               g['tipo'],
            'cantidad':           g['cantidad'],
            'habilitados':        g['habilitados'],
            'inhabilitados':      g['inhabilitados'],
            'asignados':          g['asignados'],
            'costo_mensual':      float(g['costo_mensual']),
            'costo_promedio':     float(g['costo_promedio']),
            'costo_bitdefender':  float(bd_grupo),
            'aplica_bitdefender': g['aplica_bitdefender'],
            'total':              float(g['costo_mensual'] + bd_grupo),
        })

    total_general = total_arrendamiento + total_bitdefender_global

    # Colaboradores del área (con su cantidad de dispositivos y costo)
    colaboradores_resumen = (
        asign_qs.values(
            'g216_colaborador__g215_id',
            'g216_colaborador__g215_nombre',
            'g216_colaborador__g215_documento',
        )
        .annotate(
            cantidad=Count('g216_dispositivo_id'),
            costo_arrendamiento=Sum('g216_dispositivo__g212_valor_arrendamiento'),
        )
        .order_by('g216_colaborador__g215_nombre')
    )
    colaboradores_list = [{
        'id':                    row['g216_colaborador__g215_id'],
        'nombre':                row['g216_colaborador__g215_nombre'],
        'documento':             row['g216_colaborador__g215_documento'],
        'cantidad_dispositivos': row['cantidad'],
        'costo_arrendamiento':   float(row['costo_arrendamiento'] or Decimal('0')),
    } for row in colaboradores_resumen]

    return _json_ok({
        'total':         total,
        'habilitados':   habilitados,
        'otros':         otros,
        'grupos':        grupos_list,
        'colaboradores': colaboradores_list,
        'resumen': {
            'costo_arrendamiento': float(total_arrendamiento),
            'costo_promedio':      float(total_promedio),
            'costo_bitdefender':   float(total_bitdefender_global),
            'total':               float(total_general),
        },
    })


#  INACTIVOS

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_inactivos(request):
    """
    Lista de dispositivos inactivos (Estado = Eliminado/Obsoleto/Devuelto)
    con filtros opcionales. Lee directamente de Dispositivo: un dispositivo
    inactivo sigue siendo el mismo registro, solo se excluye del Inventario
    activo mientras tenga uno de estos estados (ver api_dispositivos).
    """
    base_qs = Dispositivo.objects.filter(g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS)

    qs = base_qs.select_related('g212_tipo', 'g212_marca', 'g212_propietario', 'g212_estado', 'g212_co')

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(g212_serial__icontains=q) |
            Q(g212_propietario__g203_propietario__icontains=q) |
            Q(g212_marca__g202_marca__icontains=q)
        )

    tipo_id = request.GET.get('tipo_id')
    if tipo_id:
        qs = qs.filter(g212_tipo_id=tipo_id)

    estado_id = request.GET.get('estado_id')
    if estado_id:
        qs = qs.filter(g212_estado_id=estado_id)

    inactivos = []
    for d in qs.order_by('g212_serial'):
        inactivos.append({
            'id':          d.g212_id,
            'serial':      d.g212_serial or '—',
            'tipo':        d.g212_tipo.g200_tipo_dispositivo if d.g212_tipo else '—',
            'tipo_id':     d.g212_tipo_id,
            'marca':       d.g212_marca.g202_marca if d.g212_marca else '—',
            'marca_id':    d.g212_marca_id,
            'modelo':      d.g212_nombre_equipo or '—',
            'propietario': d.g212_propietario.g203_propietario if d.g212_propietario else '—',
            'propietario_id': d.g212_propietario_id,
            'estado':      d.g212_estado.g201_descripcion if d.g212_estado else '—',
            'estado_id':   d.g212_estado_id,
            'co':          f"{d.g212_co.g207_co} — {d.g212_co.g207_descripcion_co}" if d.g212_co else '—',
            'co_id':       d.g212_co_id,
            'observaciones': d.g212_observaciones or '',
            'fecha_registro': d.g212_fecha_registro.strftime('%d/%m/%Y'),
        })

    # Stats
    stats = {
        'total':      base_qs.count(),
        'eliminados': base_qs.filter(g212_estado__g201_descripcion='ELIMINADO').count(),
        'obsoletos':  base_qs.filter(g212_estado__g201_descripcion='OBSOLETO').count(),
        'devueltos':  base_qs.filter(g212_estado__g201_descripcion='DEVUELTO').count(),
    }

    return _json_ok({'inactivos': inactivos, 'stats': stats})

@login_required(login_url='login')
@require_http_methods(['PUT'])
def api_inactivo_editar(request, pk):
    """
    Edita un dispositivo inactivo. Es el mismo registro de Dispositivo: si
    aquí le cambian el Estado a uno activo, vuelve a aparecer en Inventario.
    """
    d = get_object_or_404(Dispositivo, pk=pk)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')

    required = ['propietario_id', 'estado_id']
    for f in required:
        if not body.get(f):
            return _json_err(f'Campo requerido: {f}')

    estado_anterior_desc = d.g212_estado.g201_descripcion if d.g212_estado else None

    d.g212_serial       = body.get('serial', d.g212_serial)
    d.g212_tipo_id      = body.get('tipo_id') or None
    d.g212_marca_id     = body.get('marca_id') or None
    d.g212_propietario_id = body['propietario_id']
    d.g212_estado_id    = body['estado_id']
    d.g212_co_id        = body.get('co_id') or None
    d.g212_observaciones = body.get('observaciones', '')
    d.save()

    responsable = request.user.get_full_name() or request.user.username
    _registrar_historial_auto(
        dispositivo    = d,
        nombre_novedad = 'EDICIÓN',
        responsable    = responsable,
        observaciones  = 'Dispositivo editado desde Inactivos.',
        co             = d.g212_co,
    )
    _registrar_transicion_inactivo(d, estado_anterior_desc, responsable, co=d.g212_co)

    return _json_ok({'id': d.g212_id})


# ═══════════════════════════════════════════════════
#  CHECKLIST DE INVENTARIO
# ═══════════════════════════════════════════════════

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_checklist_items(request):
    """
    Catálogo de preguntas del Checklist de Inventario (todas, activas e
    inactivas), agrupadas por sección y ordenadas.
      ?tipo_dispositivo_id=  si llega, además de todas las preguntas
      genéricas (sin tipo asignado) incluye las propias de ese tipo
      (ej. el checklist de PORTÁTIL).
    """
    items = ItemChecklist.objects.select_related('g236_tipo_dispositivo').all()
    tipo_id = request.GET.get('tipo_dispositivo_id')
    if tipo_id:
        items = items.filter(Q(g236_tipo_dispositivo__isnull=True) | Q(g236_tipo_dispositivo_id=tipo_id))
    return _json_ok([
        {
            'id':       i.g236_id,
            'pregunta': i.g236_pregunta,
            'activo':   i.g236_estado,
            'seccion':  i.g236_seccion,
            'orden':    i.g236_orden,
            'es_texto': i.g236_es_texto,
            'tipo_dispositivo_id':     i.g236_tipo_dispositivo_id,
            'tipo_dispositivo_nombre': i.g236_tipo_dispositivo.g200_tipo_dispositivo if i.g236_tipo_dispositivo else '',
        }
        for i in items
    ])


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_checklist_item_crear(request):
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')
    pregunta = (body.get('pregunta') or '').strip()
    if not pregunta:
        return _json_err('La pregunta es obligatoria')
    item = ItemChecklist.objects.create(
        g236_pregunta          = pregunta,
        g236_estado            = True,
        g236_seccion           = (body.get('seccion') or '').strip(),
        g236_tipo_dispositivo_id = body.get('tipo_dispositivo_id') or None,
        g236_orden             = body.get('orden') or 0,
        g236_es_texto          = bool(body.get('es_texto')),
    )
    return _json_ok({'id': item.g236_id, 'pregunta': item.g236_pregunta, 'activo': item.g236_estado})


@login_required(login_url='login')
@require_http_methods(['PUT'])
def api_checklist_item_editar(request, pk):
    item = get_object_or_404(ItemChecklist, pk=pk)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')
    pregunta = (body.get('pregunta') or '').strip()
    if pregunta:
        item.g236_pregunta = pregunta
    if 'activo' in body:
        item.g236_estado = bool(body.get('activo'))
    if 'seccion' in body:
        item.g236_seccion = (body.get('seccion') or '').strip()
    if 'tipo_dispositivo_id' in body:
        item.g236_tipo_dispositivo_id = body.get('tipo_dispositivo_id') or None
    if 'orden' in body:
        item.g236_orden = body.get('orden') or 0
    if 'es_texto' in body:
        item.g236_es_texto = bool(body.get('es_texto'))
    item.save()
    return _json_ok({'id': item.g236_id, 'pregunta': item.g236_pregunta, 'activo': item.g236_estado})


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_checklist_stats(request):
    """Resumen para el encabezado de la pantalla de Checklist."""
    total = Dispositivo.objects.exclude(g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS).count()
    con_checklist = ChecklistDispositivo.objects.filter(g237_dispositivo__isnull=False) \
        .exclude(g237_dispositivo__g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS) \
        .values('g237_dispositivo_id').distinct().count()
    return _json_ok({
        'total': total,
        'con_checklist': con_checklist,
        'pendientes': max(total - con_checklist, 0),
    })


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_checklist_tipos_disponibles(request):
    """
    Tipos de dispositivo que ya tienen preguntas de checklist configuradas
    (para que el filtro de la pantalla Checklist no deje elegir un tipo que
    todavía no tiene nada armado). Se actualiza solo a medida que se agregan
    preguntas nuevas para otros tipos desde "Administrar preguntas".
    """
    tipo_ids = ItemChecklist.objects.filter(
        g236_tipo_dispositivo__isnull=False, g236_estado=True
    ).values_list('g236_tipo_dispositivo_id', flat=True).distinct()
    tipos = TipoDispositivo.objects.filter(g200_id__in=tipo_ids).order_by('g200_tipo_dispositivo')
    return _json_ok([{'id': t.g200_id, 'nombre': t.g200_tipo_dispositivo} for t in tipos])


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_checklist_dispositivos(request):
    """
    Lista de dispositivos para la pantalla de Checklist: además de los datos
    del dispositivo, incluye el colaborador responsable (si está asignado) y
    los datos del último checklist realizado (fecha, observaciones, y si ya
    se le hizo o no), para saber de un vistazo qué falta por revisar.
    """
    qs = Dispositivo.objects.select_related('g212_tipo', 'g212_marca', 'g212_estado') \
        .exclude(g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS)

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(g212_serial__icontains=q) |
            Q(g212_propietario__g203_propietario__icontains=q) |
            Q(g212_marca__g202_marca__icontains=q)
        )
    tipo_id = request.GET.get('tipo')
    if tipo_id:
        qs = qs.filter(g212_tipo_id=tipo_id)
    estado_id = request.GET.get('estado')
    if estado_id:
        qs = qs.filter(g212_estado_id=estado_id)

    dispositivos = list(qs.order_by('g212_serial'))
    ids = [d.g212_id for d in dispositivos]

    responsables = {
        a.g216_dispositivo_id: a.g216_colaborador.g215_nombre
        for a in AsignacionColaborador.objects.filter(g216_dispositivo_id__in=ids)
                                                .select_related('g216_colaborador')
    }

    ultimos = {}
    for c in ChecklistDispositivo.objects.filter(g237_dispositivo_id__in=ids).order_by('g237_fecha'):
        ultimos[c.g237_dispositivo_id] = c  # se queda el último por orden ascendente

    resultado = []
    for d in dispositivos:
        ultimo = ultimos.get(d.g212_id)
        resultado.append({
            'id':                 d.g212_id,
            'serial':             d.g212_serial,
            'tipo':               d.g212_tipo.g200_tipo_dispositivo if d.g212_tipo else '—',
            'tipo_id':            d.g212_tipo_id,
            'estado':             d.g212_estado.g201_descripcion if d.g212_estado else '—',
            'responsable':        responsables.get(d.g212_id, ''),
            'observaciones':      ultimo.g237_observaciones if ultimo else '',
            'fecha_checklist':    _fecha_local_str(ultimo.g237_fecha) if ultimo else '',
            'checklist_realizado': bool(ultimo),
            'ultimo_checklist_id': ultimo.g237_id if ultimo else None,
        })

    return _json_ok(resultado)


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_checklist_dispositivo_guardar(request, pk):
    """
    Guarda un checklist nuevo para un dispositivo directamente desde la
    pantalla de Checklist (independiente de si cambia o no su Estado).
    """
    d = get_object_or_404(Dispositivo, pk=pk)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')
    responsable = _nombre_completo_usuario(request)
    _guardar_checklist_dispositivo(d, body, responsable)
    return _json_ok({'id': d.g212_id})


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_checklist_lista(request):
    """Historial de checklists ya realizados, con filtro opcional por serial."""
    qs = ChecklistDispositivo.objects.select_related('g237_dispositivo').order_by('-g237_fecha')

    dispositivo_id = request.GET.get('dispositivo_id')
    if dispositivo_id:
        qs = qs.filter(g237_dispositivo_id=dispositivo_id)

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(g237_dispositivo__g212_serial__icontains=q) |
            Q(g237_dispositivo_desc__icontains=q)
        )

    registros = []
    for c in qs:
        registros.append({
            'id':            c.g237_id,
            'serial':        c.g237_dispositivo.g212_serial if c.g237_dispositivo else (c.g237_dispositivo_desc or '(dispositivo eliminado)'),
            'responsable':   c.g237_responsable,
            'observaciones': c.g237_observaciones or '',
            'fecha':         _fecha_local_str(c.g237_fecha),
        })

    return _json_ok(registros)


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_checklist_detalle(request, pk):
    """Detalle completo de un checklist realizado: respuestas y firma."""
    c = get_object_or_404(ChecklistDispositivo.objects.select_related('g237_dispositivo__g212_tipo'), pk=pk)
    respuestas = [
        {
            'item_id':     r.g238_item_id,
            'pregunta':    r.g238_item_desc,
            'seccion':     r.g238_item.g236_seccion if r.g238_item and r.g238_item.g236_seccion else 'GENERAL',
            'respuesta':   r.g238_respuesta,
            'valor_texto': r.g238_valor_texto or '',
            'es_texto':    r.g238_item.g236_es_texto if r.g238_item else bool(r.g238_valor_texto),
        }
        for r in c.respuestas.all()
    ]
    asignado_a = None
    if c.g237_dispositivo:
        asign = AsignacionColaborador.objects.filter(g216_dispositivo=c.g237_dispositivo).select_related('g216_colaborador').first()
        if asign:
            asignado_a = asign.g216_colaborador.g215_nombre

    logo_b64  = ''
    logo_path = os.path.join(settings.BASE_DIR, 'index', 'static', 'img', 'imagen.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_b64 = 'data:image/png;base64,' + base64.b64encode(f.read()).decode()

    return _json_ok({
        'id':            c.g237_id,
        'dispositivo_id': c.g237_dispositivo_id,
        'tipo_dispositivo_id': c.g237_dispositivo.g212_tipo_id if c.g237_dispositivo else None,
        'tipo':          c.g237_dispositivo.g212_tipo.g200_tipo_dispositivo if (c.g237_dispositivo and c.g237_dispositivo.g212_tipo) else '—',
        'serial':        c.g237_dispositivo.g212_serial if c.g237_dispositivo else (c.g237_dispositivo_desc or '(dispositivo eliminado)'),
        'asignado_a':    asignado_a or 'Sin asignar',
        'responsable':   c.g237_responsable,
        'logo':          logo_b64,
        'observaciones': c.g237_observaciones or '',
        'resp_nombre':   c.g237_resp_nombre or '',
        'resp_cedula':   c.g237_resp_cedula or '',
        'resp_area':     c.g237_resp_area or '',
        'resp_cargo':    c.g237_resp_cargo or '',
        'fecha':         _fecha_local_str(c.g237_fecha),
        'respuestas':    respuestas,
    })


@login_required(login_url='login')
@require_http_methods(['PUT'])
def api_checklist_editar(request, pk):
    """
    Corrige un checklist ya guardado: respuestas, observaciones y datos
    del responsable. Existe para cuando quedó incompleto o con un error
    al guardarlo la primera vez, sin tener que rehacerlo desde cero.
    """
    cd = get_object_or_404(ChecklistDispositivo, pk=pk)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')

    if 'observaciones' in body:
        cd.g237_observaciones = body.get('observaciones', '')
    if 'resp_nombre' in body:
        cd.g237_resp_nombre = body.get('resp_nombre', '')
    if 'resp_cedula' in body:
        cd.g237_resp_cedula = body.get('resp_cedula', '')
    if 'resp_area' in body:
        cd.g237_resp_area = body.get('resp_area', '')
    if 'resp_cargo' in body:
        cd.g237_resp_cargo = body.get('resp_cargo', '')
    cd.save()

    for r in body.get('respuestas', []):
        item_id = r.get('item_id')
        if not item_id:
            continue
        item = ItemChecklist.objects.filter(pk=item_id).first()
        RespuestaChecklist.objects.update_or_create(
            g238_checklist=cd, g238_item_id=item_id,
            defaults={
                'g238_item_desc':   item.g236_pregunta if item else (r.get('pregunta') or ''),
                'g238_respuesta':   bool(r.get('respuesta')),
                'g238_valor_texto': r.get('valor_texto') or None,
            }
        )

    if cd.g237_dispositivo:
        _registrar_historial_auto(
            dispositivo    = cd.g237_dispositivo,
            nombre_novedad = 'CHECKLIST CORREGIDO',
            responsable    = _nombre_completo_usuario(request),
            observaciones  = 'Se corrigió/completó el checklist de inventario guardado previamente.',
            co             = cd.g237_dispositivo.g212_co,
        )

    return _json_ok({'id': cd.g237_id})


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_checklist_colaborador_buscar(request):
    """
    Búsqueda liviana de colaboradores por cédula, para autocompletar
    "Datos del Responsable" (Nombre / Cédula / Área / Cargo) del Checklist.
    """
    q = request.GET.get('q', '').strip()
    qs = Colaborador.objects.select_related('g215_Area')
    if q:
        qs = qs.filter(g215_documento__icontains=q)
    qs = qs.order_by('g215_documento')[:15]
    return _json_ok([
        {
            'id':        c.g215_id,
            'nombre':    c.g215_nombre,
            'documento': c.g215_documento,
            'area':      c.g215_Area.g228_nombre if c.g215_Area else '',
            'cargo':     c.g215_cargo,
        }
        for c in qs
    ])


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_checklist_mi_responsable(request):
    """
    El login del Dashboard usa la cédula como username, así que casi siempre
    coincide con un registro de Colaborador. Se usa para precargar "Datos
    del Responsable" con los datos del usuario que inició sesión, sin que
    tenga que buscarse a sí mismo. Si no hay coincidencia, devuelve null.
    """
    c = Colaborador.objects.select_related('g215_Area').filter(g215_documento=request.user.username).first()
    if not c:
        return _json_ok(None)
    return _json_ok({
        'id':        c.g215_id,
        'nombre':    c.g215_nombre,
        'documento': c.g215_documento,
        'area':      c.g215_Area.g228_nombre if c.g215_Area else '',
        'cargo':     c.g215_cargo,
    })


@login_required(login_url='login')
@require_http_methods(['GET'])
@xframe_options_sameorigin
def api_checklist_pdf(request, pk):
    """Genera el PDF de un checklist ya guardado, con el mismo estilo que las Actas."""
    cd = get_object_or_404(
        ChecklistDispositivo.objects.select_related('g237_dispositivo__g212_tipo', 'g237_dispositivo__g212_co'),
        pk=pk
    )
    respuestas = list(cd.respuestas.all().order_by('g238_id'))

    colaborador = None
    if cd.g237_dispositivo:
        asign = AsignacionColaborador.objects.filter(g216_dispositivo=cd.g237_dispositivo).select_related('g216_colaborador').first()
        if asign:
            colaborador = asign.g216_colaborador

    logo_b64  = ''
    logo_path = os.path.join(settings.BASE_DIR, 'index', 'static', 'img', 'imagen.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_b64 = 'data:image/png;base64,' + base64.b64encode(f.read()).decode()

    html_checklist = _construir_html_checklist(cd, respuestas, colaborador, logo_b64)
    buffer = BytesIO()
    pisa.CreatePDF(html_checklist, dest=buffer)
    pdf_bytes = buffer.getvalue()

    serial = cd.g237_dispositivo.g212_serial if cd.g237_dispositivo else (cd.g237_dispositivo_desc or 'dispositivo')
    nombre_archivo = f"Checklist_{serial}_{cd.g237_id}.pdf"
    disposicion = 'inline' if request.GET.get('inline') else 'attachment'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'{disposicion}; filename="{nombre_archivo}"'
    return response


def _construir_html_checklist(cd, respuestas, colaborador, logo_b64):
    """Arma el HTML del PDF del Checklist de Inventario, agrupado por sección
    (mismo formato que la plantilla en Excel que ya se usaba manualmente)."""
    d = cd.g237_dispositivo
    serial = d.g212_serial if d else (cd.g237_dispositivo_desc or '—')
    tipo   = d.g212_tipo.g200_tipo_dispositivo if d and d.g212_tipo else '—'

    # Solo "Controlador de Dominio" usa la columna "Observación" — las demás
    # secciones usan "Registro" (mismo criterio que en la pantalla).
    columna_por_seccion = {'CONTROLADOR DE DOMINIO': 'OBSERVACIÓN'}

    # Cada sección arma su PROPIA tabla completa (encabezado + filas), en vez de
    # compartir una sola tabla larga con filas de colspan intercaladas: xhtml2pdf
    # calcula mal el ancho de columnas cuando una tabla mezcla filas con colspan
    # y filas normales, y las columnas SI/NO terminaban aplastadas/cortadas.
    def _tabla_seccion(seccion, filas_seccion):
        columna = columna_por_seccion.get(seccion, 'REGISTRO')
        return f"""
        <div style="background:#1e3a5f;color:#ffffff;padding:3px 8px;font-weight:bold;font-size:9.5px;border:1px solid #1e3a5f">{seccion}</div>
        <table style="width:460pt;border-collapse:collapse;margin-bottom:5px;font-size:9px">
          <colgroup>
            <col style="width:190pt"/>
            <col style="width:200pt"/>
            <col style="width:60pt"/>
          </colgroup>
          <thead>
            <tr style="background:#e5edf5">
              <th style="padding:2px 8px;border:1px solid #cbd5e1;text-align:left">ITEM</th>
              <th style="padding:2px 8px;border:1px solid #cbd5e1;text-align:left">{columna}</th>
              <th style="padding:2px 6px;border:1px solid #cbd5e1;text-align:center">RESULTADO</th>
            </tr>
          </thead>
          <tbody>{filas_seccion}</tbody>
        </table>"""

    secciones_html = ''
    seccion_actual = None
    filas_seccion = ''
    for r in respuestas:
        seccion = r.g238_item.g236_seccion if r.g238_item and r.g238_item.g236_seccion else 'GENERAL'
        if seccion != seccion_actual:
            if seccion_actual is not None:
                secciones_html += _tabla_seccion(seccion_actual, filas_seccion)
            seccion_actual = seccion
            filas_seccion = ''

        if r.g238_respuesta:
            resultado = '<span style="color:#15803d">&#10003; SI</span>'
        else:
            resultado = '<span style="color:#b91c1c">X NO</span>'
        filas_seccion += f"""
        <tr>
          <td style="padding:2px 8px;border:1px solid #e5e7eb">{r.g238_item_desc}</td>
          <td style="padding:2px 8px;border:1px solid #e5e7eb;font-size:9px">{r.g238_valor_texto or '&nbsp;'}</td>
          <td style="padding:2px 6px;border:1px solid #e5e7eb;text-align:center;font-weight:bold">{resultado}</td>
        </tr>"""

    if seccion_actual is not None:
        secciones_html += _tabla_seccion(seccion_actual, filas_seccion)

    if not secciones_html:
        secciones_html = '<p style="text-align:center;color:#6b7280">Sin respuestas registradas</p>'

    logo_html  = f'<img src="{logo_b64}" style="max-height:85px;max-width:130px"/>' if logo_b64 else '<b style="font-size:16px;color:#1e3a5f">AM&amp;M</b>'

    resp_nombre = cd.g237_resp_nombre or ''
    resp_cedula = cd.g237_resp_cedula or ''
    resp_area   = cd.g237_resp_area or ''
    resp_cargo  = cd.g237_resp_cargo or ''

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<style>
  body {{ margin:0; padding:0; font-family: Arial, sans-serif; font-size: 11px; color: #111; }}
  @page {{ margin: 10mm 12mm 10mm 12mm; }}
  table {{ border-collapse: collapse; width: 100%; }}
  td, th {{ font-size: 11px; }}
</style>
</head>
<body>

<table style="width:100%;margin-bottom:6px;border-bottom:2px solid #111;padding-bottom:6px">
  <tr>
    <td style="width:140px;vertical-align:middle">{logo_html}</td>
    <td style="text-align:center;vertical-align:middle;padding:0 10px">
      <div style="font-size:13px;font-weight:bold;text-transform:uppercase">CHECKLIST DE INGRESO Y EGRESO DE EQUIPOS</div>
      <div style="font-size:11px;font-weight:bold;margin-top:3px">GESTIÓN DE TECNOLOGÍA DE LA INFORMACIÓN Y LA COMUNICACIÓN</div>
    </td>
    <td style="width:110px"></td>
  </tr>
</table>

<table style="width:100%;margin-bottom:10px;font-size:11px">
  <tr>
    <td style="width:130px;padding:2px 0"><b>FECHA:</b></td>
    <td style="padding:2px 0">{_fecha_local_str(cd.g237_fecha)}</td>
    <td style="width:130px;padding:2px 0"><b>SERIAL:</b></td>
    <td style="padding:2px 0">{serial}</td>
  </tr>
  <tr>
    <td style="padding:2px 0"><b>TIPO DE EQUIPO:</b></td>
    <td style="padding:2px 0">{tipo}</td>
    <td style="padding:2px 0"><b>REGISTRADO POR:</b></td>
    <td style="padding:2px 0">{cd.g237_responsable}</td>
  </tr>
  <tr>
    <td style="padding:2px 0"><b>ASIGNADO A:</b></td>
    <td colspan="3" style="padding:2px 0">{colaborador.g215_nombre if colaborador else 'Sin asignar'}</td>
  </tr>
</table>

{secciones_html}

<table style="width:100%;margin-top:10px;font-size:10.5px">
  <tr style="background:#1e3a5f;color:#ffffff">
    <td colspan="4" style="padding:5px 8px;border:1px solid #1e3a5f;font-weight:bold;text-align:center">DATOS DEL RESPONSABLE</td>
  </tr>
  <tr>
    <td style="padding:4px 8px;border:1px solid #cbd5e1;background:#e5edf5;font-weight:bold;width:20%">NOMBRE COMPLETO</td>
    <td colspan="3" style="padding:4px 8px;border:1px solid #cbd5e1">{resp_nombre or '&nbsp;'}</td>
  </tr>
  <tr>
    <td style="padding:4px 8px;border:1px solid #cbd5e1;background:#e5edf5;font-weight:bold">CEDULA</td>
    <td colspan="3" style="padding:4px 8px;border:1px solid #cbd5e1">{resp_cedula or '&nbsp;'}</td>
  </tr>
  <tr>
    <td style="padding:4px 8px;border:1px solid #cbd5e1;background:#e5edf5;font-weight:bold">AREA</td>
    <td style="padding:4px 8px;border:1px solid #cbd5e1;width:30%">{resp_area or '&nbsp;'}</td>
    <td style="padding:4px 8px;border:1px solid #cbd5e1;background:#e5edf5;font-weight:bold;width:20%">CARGO</td>
    <td style="padding:4px 8px;border:1px solid #cbd5e1">{resp_cargo or '&nbsp;'}</td>
  </tr>
</table>

<table style="width:100%;margin-top:10px;font-size:10.5px">
  <tr style="background:#1e3a5f;color:#ffffff">
    <td style="padding:5px 8px;border:1px solid #1e3a5f;font-weight:bold">OBSERVACIONES:</td>
  </tr>
  <tr>
    <td style="padding:10px 8px;border:1px solid #cbd5e1;min-height:60px;height:60px;vertical-align:top">{cd.g237_observaciones or '&nbsp;'}</td>
  </tr>
</table>

</body>
</html>"""


# ═══════════════════════════════════════════════════════════════
# NOVEDADES GENERALES — bitácora independiente (no se liga a
# dispositivos ni colaboradores). Catálogo de tipos + campos
# dinámicos por tipo, igual patrón que Checklist de Inventario.
# ═══════════════════════════════════════════════════════════════

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_novedades_tipos(request):
    """Catálogo de tipos de novedad (todos, activos e inactivos)."""
    tipos = TipoNovedadGeneral.objects.all()
    return _json_ok([
        {'id': t.g239_id, 'nombre': t.g239_nombre, 'activo': t.g239_estado}
        for t in tipos
    ])


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_novedades_tipo_crear(request):
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')
    nombre = (body.get('nombre') or '').strip()
    if not nombre:
        return _json_err('El nombre del tipo es obligatorio')
    tipo = TipoNovedadGeneral.objects.create(g239_nombre=nombre, g239_estado=True)
    return _json_ok({'id': tipo.g239_id, 'nombre': tipo.g239_nombre, 'activo': tipo.g239_estado})


@login_required(login_url='login')
@require_http_methods(['PUT'])
def api_novedades_tipo_editar(request, pk):
    tipo = get_object_or_404(TipoNovedadGeneral, pk=pk)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')
    nombre = (body.get('nombre') or '').strip()
    if nombre:
        tipo.g239_nombre = nombre
    if 'activo' in body:
        tipo.g239_estado = bool(body.get('activo'))
    tipo.save()
    return _json_ok({'id': tipo.g239_id, 'nombre': tipo.g239_nombre, 'activo': tipo.g239_estado})


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_novedades_campos(request):
    """
    Campos configurados para un tipo de novedad.
      ?tipo_id=      obligatorio.
      ?todos=1       incluye también los campos inactivos (para
                     administrarlos); sin esto, solo devuelve los activos
                     (para armar el formulario de "Registrar novedad").
    """
    tipo_id = request.GET.get('tipo_id')
    if not tipo_id:
        return _json_err('Falta tipo_id')
    campos = CampoNovedadGeneral.objects.filter(g240_tipo_novedad_id=tipo_id)
    if not request.GET.get('todos'):
        campos = campos.filter(g240_estado=True)
    return _json_ok([
        {
            'id':     c.g240_id,
            'nombre': c.g240_nombre_campo,
            'orden':  c.g240_orden,
            'activo': c.g240_estado,
            'tipo_id': c.g240_tipo_novedad_id,
        }
        for c in campos
    ])


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_novedades_campo_crear(request):
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')
    tipo_id = body.get('tipo_id')
    nombre_campo = (body.get('nombre_campo') or '').strip()
    if not tipo_id or not nombre_campo:
        return _json_err('El tipo y el nombre del campo son obligatorios')
    campo = CampoNovedadGeneral.objects.create(
        g240_tipo_novedad_id=tipo_id,
        g240_nombre_campo=nombre_campo,
        g240_orden=body.get('orden') or 0,
        g240_estado=True,
    )
    return _json_ok({'id': campo.g240_id, 'nombre': campo.g240_nombre_campo, 'activo': campo.g240_estado})


@login_required(login_url='login')
@require_http_methods(['PUT'])
def api_novedades_campo_editar(request, pk):
    campo = get_object_or_404(CampoNovedadGeneral, pk=pk)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')
    nombre_campo = (body.get('nombre_campo') or '').strip()
    if nombre_campo:
        campo.g240_nombre_campo = nombre_campo
    if 'activo' in body:
        campo.g240_estado = bool(body.get('activo'))
    if 'orden' in body:
        campo.g240_orden = body.get('orden') or 0
    campo.save()
    return _json_ok({'id': campo.g240_id, 'nombre': campo.g240_nombre_campo, 'activo': campo.g240_estado})


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_novedades_lista(request):
    """
    Lista de novedades registradas, con filtros opcionales:
      ?q=          busca en tipo, responsable y observaciones.
      ?tipo_id=
      ?fecha_desde=  ?fecha_hasta=   (formato YYYY-MM-DD)
    Paginación es del lado del frontend, igual que Inventario.
    """
    qs = NovedadGeneral.objects.select_related('g241_tipo').prefetch_related('respuestas')

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(g241_tipo_desc__icontains=q) |
            Q(g241_responsable__icontains=q) |
            Q(g241_observaciones__icontains=q)
        )
    tipo_id = request.GET.get('tipo_id')
    if tipo_id:
        qs = qs.filter(g241_tipo_id=tipo_id)
    fecha_desde = request.GET.get('fecha_desde')
    if fecha_desde:
        qs = qs.filter(g241_fecha__date__gte=fecha_desde)
    fecha_hasta = request.GET.get('fecha_hasta')
    if fecha_hasta:
        qs = qs.filter(g241_fecha__date__lte=fecha_hasta)

    registros = []
    for n in qs:
        detalle = ' · '.join(
            f"{r.g242_campo_desc}: {r.g242_observacion}"
            for r in n.respuestas.all() if r.g242_observacion
        )
        registros.append({
            'id':            n.g241_id,
            'tipo_id':       n.g241_tipo_id,
            'tipo':          n.g241_tipo_desc or (n.g241_tipo.g239_nombre if n.g241_tipo else '—'),
            'responsable':   n.g241_responsable,
            'observaciones': n.g241_observaciones or '',
            'detalle':       detalle,
            'fecha':         _fecha_local_str(n.g241_fecha),
        })
    return _json_ok(registros)


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_novedades_guardar(request):
    """Crea un registro nuevo de Novedad General con sus campos."""
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')
    tipo_id = body.get('tipo_id')
    if not tipo_id:
        return _json_err('Debes elegir un tipo de novedad')
    tipo = get_object_or_404(TipoNovedadGeneral, pk=tipo_id)

    novedad = NovedadGeneral.objects.create(
        g241_tipo          = tipo,
        g241_tipo_desc      = tipo.g239_nombre,
        g241_responsable    = _nombre_completo_usuario(request),
        g241_observaciones  = body.get('observaciones', ''),
    )
    for c in body.get('campos', []):
        campo_id = c.get('campo_id')
        campo = CampoNovedadGeneral.objects.filter(pk=campo_id).first()
        RespuestaNovedadGeneral.objects.create(
            g242_novedad     = novedad,
            g242_campo       = campo,
            g242_campo_desc  = campo.g240_nombre_campo if campo else (c.get('nombre') or ''),
            g242_observacion = c.get('observacion') or '',
        )
    return _json_ok({'id': novedad.g241_id})


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_novedades_detalle(request, pk):
    n = get_object_or_404(NovedadGeneral.objects.select_related('g241_tipo'), pk=pk)
    respuestas = [{
        'campo_id':    r.g242_campo_id,
        'campo':       r.g242_campo_desc,
        'observacion': r.g242_observacion or '',
    } for r in n.respuestas.all()]
    return _json_ok({
        'id':            n.g241_id,
        'tipo_id':       n.g241_tipo_id,
        'tipo':          n.g241_tipo_desc or (n.g241_tipo.g239_nombre if n.g241_tipo else '—'),
        'responsable':   n.g241_responsable,
        'observaciones': n.g241_observaciones or '',
        'fecha':         _fecha_local_str(n.g241_fecha),
        'respuestas':    respuestas,
    })


# COLABORADORES

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_colaboradores(request):
    """Lista de colaboradores con paginación en el backend."""
    qs = Colaborador.objects.select_related('g215_co', 'g215_estado')

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(g215_documento__icontains=q) |
            Q(g215_nombre__icontains=q) |
            Q(g215_cargo__icontains=q)
        )

    # Orden solicitado por el paginador de la tabla (clic en encabezado)
    SORT_FIELDS_COLAB = {
        'documento': 'g215_documento',
        'nombre':    'g215_nombre',
        'co':        'g215_co__g207_co',
        'cargo':     'g215_cargo',
        'estado':    'g215_estado__g201_descripcion',
    }
    sort_field = SORT_FIELDS_COLAB.get(request.GET.get('sort', ''), 'g215_nombre')
    if request.GET.get('dir', 'asc') == 'desc':
        sort_field = f'-{sort_field}'
    qs = qs.order_by(sort_field)
    total = qs.count()

    # Paginación
    try:
        page      = max(1, int(request.GET.get('page', 1)))
        page_size = min(100, max(10, int(request.GET.get('page_size', 25))))
    except (ValueError, TypeError):
        page, page_size = 1, 25

    offset = (page - 1) * page_size
    qs_page = qs[offset: offset + page_size]

    # IDs de la página para obtener asignaciones y actas de golpe
    ids_page = [c.g215_id for c in qs_page]

    # Asignaciones de todos los colaboradores de la página en 1 query
    asig_map = {}
    for a in AsignacionColaborador.objects.filter(
        g216_colaborador_id__in=ids_page
    ).select_related('g216_dispositivo__g212_tipo', 'g216_dispositivo__g212_marca'):
        asig_map.setdefault(a.g216_colaborador_id, []).append({
            'id':     a.g216_dispositivo.g212_id,
            'tipo':   a.g216_dispositivo.g212_tipo.g200_tipo_dispositivo if a.g216_dispositivo.g212_tipo else '—',
            'marca':  a.g216_dispositivo.g212_marca.g202_marca if a.g216_dispositivo.g212_marca else '—',
            'serial': a.g216_dispositivo.g212_serial,
        })

    # Actas de todos los colaboradores de la página en 1 query
    actas_map = {}
    for a in Acta.objects.filter(g217_colaborador_id__in=ids_page).order_by('-g217_fecha'):
        actas_map.setdefault(a.g217_colaborador_id, []).append({
            'id':      a.g217_id,
            'tipo':    a.g217_tipo,
            'proceso': a.g217_proceso,
            'fecha':   a.g217_fecha.strftime('%d/%m/%Y %H:%M') if a.g217_fecha else '—',
        })

    colaboradores = []
    for c in qs_page:
        colaboradores.append({
            'id':           c.g215_id,
            'documento':    c.g215_documento,
            'nombre':       c.g215_nombre,
            'co':           f"{c.g215_co.g207_co} — {c.g215_co.g207_descripcion_co}" if c.g215_co else '—',
            'co_id':        c.g215_co_id,
            'cargo':        c.g215_cargo,
            'estado':       c.g215_estado.g201_descripcion if c.g215_estado else '—',
            'estado_id':    c.g215_estado_id,
            'correo':       c.g215_correo or '',   
            'dispositivos': asig_map.get(c.g215_id, []),
            'actas':        actas_map.get(c.g215_id, []),
            
        })

    return _json_ok({
        'colaboradores': colaboradores,
        'total':         total,
        'page':          page,
        'page_size':     page_size,
        'total_pages':   (total + page_size - 1) // page_size,
    })


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_colaboradores_cargos(request):
    """Cargos distintos ya registrados en Colaborador, para el buscador del modal de creación."""
    cargos = (
        Colaborador.objects
        .exclude(g215_cargo__isnull=True)
        .exclude(g215_cargo__exact='')
        .order_by('g215_cargo')
        .values_list('g215_cargo', flat=True)
        .distinct()
    )
    data = [{'id': c, 'nombre': c} for c in cargos]
    return _json_ok(data)


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_colaborador_crear(request):
    """Crea un nuevo colaborador desde la plataforma (fuera del panel de admin)."""
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')

    documento = (body.get('documento') or '').strip()
    nombre    = (body.get('nombre') or '').strip().upper()
    cargo     = (body.get('cargo') or '').strip().upper()
    correo    = (body.get('correo') or '').strip()
    co_id     = body.get('co_id') or None
    estado_id = body.get('estado_id') or None
    area_id   = body.get('area_id') or None

    if not documento or not nombre or not cargo or not estado_id:
        return _json_err('Completa los campos requeridos: documento, nombre, cargo y estado')

    if Colaborador.objects.filter(g215_documento=documento).exists():
        return _json_err(f'Ya existe un colaborador con el documento "{documento}"')

    
    # Validamos que las FK realmente existan en ESTA base de datos.
    # Evita guardar un ID "huérfano" (p.ej. de otra base/entorno) que luego
    # se muestra vacío en la tabla sin ningún error visible.
    if co_id and not CentroOperaciones.objects.filter(pk=co_id).exists():
        return _json_err(
            'El Centro de Operación seleccionado no existe en esta base de datos. '
            'Recarga la página (Ctrl+Shift+R) para actualizar el catálogo e inténtalo de nuevo.'
        )
    if area_id and not CentroCosto.objects.filter(pk=area_id).exists():
        return _json_err(
            'El Área/Centro de Costo seleccionado no existe en esta base de datos. '
            'Recarga la página (Ctrl+Shift+R) para actualizar el catálogo e inténtalo de nuevo.'
        )
    if not Estado.objects.filter(pk=estado_id).exists():
        return _json_err(
            'El Estado seleccionado no existe en esta base de datos. '
            'Recarga la página (Ctrl+Shift+R) para actualizar el catálogo e inténtalo de nuevo.'
        )
     

    try:
        with transaction.atomic():
            c = Colaborador.objects.create(
                g215_documento = documento,
                g215_nombre    = nombre,
                g215_cargo     = cargo,
                g215_correo    = correo or None,
                g215_co_id     = co_id,
                g215_estado_id = estado_id,
                g215_Area_id   = area_id,
            )
    except Exception as e:
        return _json_err(str(e))

    return _json_ok({
        'id':        c.g215_id,
        'documento': c.g215_documento,
        'nombre':    c.g215_nombre,
    })


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_asignacion_guardar(request, colaborador_id):
    """
    Guarda la asignación de dispositivos a un colaborador.
    body = {'dispositivos': [id1, id2, ...]}
    Comportamiento: los IDs enviados reemplazan la lista completa vigente
    SOLO si se envía el parámetro 'reemplazar': true. Por defecto (false),
    se AGREGAN a las asignaciones existentes sin eliminar las anteriores.
    Así, al agregar una nueva asignación las viejas se conservan.

    Cambios de estado automáticos:
      - Dispositivo asignado  → estado ASIGNADO (id=3)
      - Colaborador con ≥1 dispositivo → estado DISPOSITIVOS OTORGADOS (id=6)
      - En modo reemplazar, dispositivos removidos → estado HABILITADO (id=1)
    """
    #  IDs de estado (tabla j201_estado) 
    ESTADO_DISP_LIBRE       = 1   # HABILITADO  (dispositivo libre)
    ESTADO_DISP_ASIGNADO    = 3   # ASIGNADO    (dispositivo en uso)
    ESTADO_COLAB_CON_DISP   = 6   # DISPOSITIVOS OTORGADOS
    ESTADO_COLAB_SIN_DISP   = 10  # SIN ASIGNACIONES
    

    c = get_object_or_404(Colaborador, pk=colaborador_id)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')

    dispositivo_ids = body.get('dispositivos', [])
    reemplazar = body.get('reemplazar', False)

    with transaction.atomic():
        if reemplazar:
            # Guardar IDs que se van a quitar para revertir su estado
            ids_anteriores = set(
                AsignacionColaborador.objects.filter(g216_colaborador=c)
                .values_list('g216_dispositivo_id', flat=True)
            )
            ids_nuevos = set(dispositivo_ids)
            ids_removidos = ids_anteriores - ids_nuevos

            # Eliminar asignaciones anteriores
            AsignacionColaborador.objects.filter(g216_colaborador=c).delete()
            # Historial automático — desasignación de los removidos
            for rid in ids_removidos:
                dev_rem = Dispositivo.objects.filter(pk=rid).first()
                if dev_rem:
                    _registrar_historial_auto(
                        dispositivo   = dev_rem,
                        nombre_novedad = 'DESASIGNACIÓN DE COLABORADOR',
                        responsable   = request.user.get_full_name() or request.user.username,
                        observaciones = f'Removido de: {c.g215_nombre}',
                        co            = dev_rem.g212_co,
                    )

            # Revertir estado de los dispositivos removidos → HABILITADO
            if ids_removidos:
                Dispositivo.objects.filter(pk__in=ids_removidos).update(
                    g212_estado_id=ESTADO_DISP_LIBRE
                )

            # Crear nuevas asignaciones y marcar dispositivos como ASIGNADO
            for dev_id in dispositivo_ids:
                dev = Dispositivo.objects.filter(pk=dev_id).first()
                if dev:
                    AsignacionColaborador.objects.create(
                        g216_colaborador=c,
                        g216_dispositivo_id=dev_id,
                    )
                    dev.g212_estado_id = ESTADO_DISP_ASIGNADO
                    dev.save(update_fields=['g212_estado_id'])
                     # Historial automático — asignación
                    _registrar_historial_auto(
                        dispositivo   = dev,
                        nombre_novedad = 'ASIGNACIÓN A COLABORADOR',
                        responsable   = request.user.get_full_name() or request.user.username,
                        observaciones = f'Asignado a: {c.g215_nombre}',
                        co            = dev.g212_co,
                    )
        else:
            # Modo acumulativo (default): agrega solo los nuevos, conserva los anteriores
            ya_asignados = set(
                AsignacionColaborador.objects.filter(g216_colaborador=c)
                .values_list('g216_dispositivo_id', flat=True)
            )
            for dev_id in dispositivo_ids:
                if dev_id not in ya_asignados:
                    dev = Dispositivo.objects.filter(pk=dev_id).first()
                    if dev:
                        # Bloquear si está asignado a OTRO colaborador
                        conflicto = AsignacionColaborador.objects.filter(
                            g216_dispositivo_id=dev_id
                        ).exclude(g216_colaborador=c).select_related('g216_colaborador').first()
                        if conflicto:
                            return _json_err(
                                f'El serial {dev.g212_serial} ya está asignado a '
                                f'{conflicto.g216_colaborador.g215_nombre}. '
                                f'Debes desasignarlo primero.'
                            )
                        AsignacionColaborador.objects.create(
                            g216_colaborador=c,
                            g216_dispositivo_id=dev_id,
                        )
                        ya_asignados.add(dev_id)
                        # Marcar dispositivo como ASIGNADO
                        dev.g212_estado_id = ESTADO_DISP_ASIGNADO
                        dev.save(update_fields=['g212_estado_id'])
                        
                        
                        # Historial automático
                        _registrar_historial_auto(
                            dispositivo   = dev,
                            nombre_novedad = 'ASIGNACIÓN A COLABORADOR',
                            responsable   = request.user.get_full_name() or request.user.username,
                            observaciones = f'Asignado a: {c.g215_nombre}',
                            co            = dev.g212_co,
                        )
                        
                        
                        

        # Actualizar estado del colaborador según si tiene dispositivos o no
        total_asignados = AsignacionColaborador.objects.filter(g216_colaborador=c).count()
        c.g215_estado_id = ESTADO_COLAB_CON_DISP if total_asignados > 0 else ESTADO_COLAB_SIN_DISP
        c.save(update_fields=['g215_estado_id'])

    return _json_ok({
        'colaborador_id': c.g215_id,
        'asignados': total_asignados,
    })
    
def _construir_html_acta(acta, colaborador, dispositivos, logo_b64):
    from datetime import datetime
    fecha_str = acta.g217_fecha.strftime('%d/%m/%Y %H:%M') if acta.g217_fecha else datetime.now().strftime('%d/%m/%Y %H:%M')

    tipo_upper = (acta.g217_tipo or '').upper()
    titulo_acta = 'ACTA DE DEVOLUCIÓN DE EQUIPOS TECNOLÓGICOS' if 'DEVOLU' in tipo_upper else 'ACTA DE ENTREGA DE EQUIPOS TECNOLÓGICOS'

    filas_dispositivos = ''
    for i, d in enumerate(dispositivos):
        carac_items = ''.join(
            f'<b>{k}:</b> {v}<br/>'
            for k, v in (d['caracteristicas'] or {}).items()
        )
        bg = '#ffffff' if i % 2 == 0 else '#f8fafc'
        filas_dispositivos += f"""
        <tr style="background:{bg}">
          <td style="padding:4px 6px;border:1px solid #e5e7eb;text-align:center;font-weight:bold">{i+1}</td>
          <td style="padding:4px 6px;border:1px solid #e5e7eb;font-weight:bold">{d['tipo']}</td>
          <td style="padding:4px 6px;border:1px solid #e5e7eb;font-size:10px">{d['serial']}</td>
          <td style="padding:4px 6px;border:1px solid #e5e7eb;font-size:10px;line-height:1.4">{carac_items}</td>
        </tr>"""

    if not filas_dispositivos:
        filas_dispositivos = '<tr><td colspan="4" style="padding:8px;text-align:center;color:#6b7280">Sin dispositivos asignados</td></tr>'

    firma_recibe_html  = f'<img src="{acta.g217_firma_recibe}" style="max-width:180px;max-height:60px"/>'  if acta.g217_firma_recibe  else '&nbsp;'
    firma_entrega_html = f'<img src="{acta.g217_firma_entrega}" style="max-width:180px;max-height:60px"/>' if acta.g217_firma_entrega else '&nbsp;'

    logo_html = f'<img src="{logo_b64}" style="max-height:70px;max-width:120px"/>' if logo_b64 else '<b style="font-size:16px;color:#1e3a5f">AM&amp;M</b>'

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<style>
  body {{
    margin: 0;
    padding: 0;
    font-family: Arial, sans-serif;
    font-size: 11px;
    color: #111;
  }}
  @page {{
    margin: 10mm 12mm 10mm 12mm;
  }}
  table {{ border-collapse: collapse; width: 100%; }}
  td, th {{ font-size: 11px; }}
</style>
</head>
<body>

<!-- ENCABEZADO -->
<table style="width:100%;margin-bottom:6px;border-bottom:2px solid #111;padding-bottom:6px">
  <tr>
    <td style="width:140px;vertical-align:middle">{logo_html}</td>
    <td style="text-align:center;vertical-align:middle;padding:0 10px">
      <div style="font-size:13px;font-weight:bold;text-transform:uppercase">{titulo_acta}</div>
      <div style="font-size:11px;font-weight:bold;margin-top:3px">GESTIÓN DE TECNOLOGÍA DE LA INFORMACIÓN Y LA COMUNICACIÓN</div>
    </td>
    <td style="width:110px;text-align:right;vertical-align:middle;font-size:10px;font-weight:bold;line-height:1.6">
      CÓDIGO: TIC-INF-F-2<br/>VERSIÓN: 6
    </td>
  </tr>
</table>

<!-- INFO COLABORADOR -->
<table style="width:100%;margin-bottom:6px;font-size:11px">
  <tr>
    <td style="width:180px;padding:2px 0"><b>FECHA:</b></td>
    <td style="padding:2px 0">{fecha_str}</td>
  </tr>
  <tr>
    <td style="padding:2px 0"><b>NOMBRE COLABORADOR:</b></td>
    <td style="padding:2px 0">{colaborador.g215_nombre}</td>
  </tr>
  <tr>
    <td style="padding:2px 0"><b>CARGO COLABORADOR:</b></td>
    <td style="padding:2px 0">{colaborador.g215_cargo}</td>
  </tr>
  <tr>
    <td style="padding:2px 0"><b>PROCESO/ÁREA COLABORADOR:</b></td>
    <td style="padding:2px 0">{acta.g217_proceso}</td>
  </tr>
</table>

<!-- TABLA DISPOSITIVOS -->
<table style="width:100%;margin-bottom:6px;font-size:10px">
  <thead>
    <tr style="background:#1e3a5f;color:#ffffff">
      <th style="padding:6px 8px;border:1px solid #1e3a5f;text-align:center;width:25px">#</th>
      <th style="padding:6px 8px;border:1px solid #1e3a5f;text-align:left;width:110px">TIPO DISPOSITIVO</th>
      <th style="padding:6px 8px;border:1px solid #1e3a5f;text-align:left;width:100px">SERIAL</th>
      <th style="padding:6px 8px;border:1px solid #1e3a5f;text-align:left">CARACTERÍSTICAS</th>
    </tr>
  </thead>
  <tbody>{filas_dispositivos}</tbody>
</table>

<!-- TEXTO LEGAL -->
<div style="font-size:9px;color:#222;text-align:justify;line-height:1.5;margin-bottom:8px;border-top:1px solid #ccc;padding-top:6px">
Certifico que los elementos detallados en el presente documento, me han sido entregados en las condiciones descritas y en buenas condiciones, operativas, funcionales y físicas para mi cuidado y custodia con el propósito de cumplir con las tareas y asignaciones propias de mi cargo en la empresa, siendo estas de mi única y exclusiva responsabilidad. Si la parte o equipo tecnológico presentase fallas o mal funcionamiento reportarlo al área de sistemas en un tiempo no mayor a 30 días para el trámite de las garantías correspondientes si las cubriese. Me comprometo a usar correctamente los recursos, y solo para los fines establecidos, a no instalar ni permitir la instalación de software para uso personal ajeno al personal de Gestión de Tecnología e Informática. Todo daño físico causado por maltrato o por el uso inapropiado de los equipos asignados y de los planes corporativos el robo o pérdida de éstos es de mi única y exclusiva responsabilidad, por lo cual autorizo el descuento del valor correspondiente del pago de nómina; así mismo al finalizar mi contrato laboral me comprometo a realizar la devolución a la totalidad de los equipos asignados y autorizo el descuento de salarios, prestaciones sociales, vacaciones, indemnizaciones, bonificaciones, auxilios y demás derechos que me correspondan el valor correspondiente a daños, pérdida o robo de los equipos en mención.
<br/><br/>
De igual manera, certifico que con el equipo tecnológico recibido daré buen uso a los recursos informáticos, conforme lo establecido en el documento TI-P-005 Política uso de recursos informáticos.
</div>

<!-- FIRMAS -->
<table style="width:100%;margin-top:8px">
  <tr>
    <td style="width:50%;text-align:center;padding-right:20px;vertical-align:bottom">
      {firma_recibe_html}
      <hr style="border:none;border-top:1px solid #333;margin:4px 0"/>
      <div style="font-weight:bold;font-size:11px">{colaborador.g215_nombre}</div>
      <div style="font-size:10px;color:#555">FIRMA QUIEN RECIBE</div>
    </td>
    <td style="width:50%;text-align:center;padding-left:20px;vertical-align:bottom">
      {firma_entrega_html}
      <hr style="border:none;border-top:1px solid #333;margin:4px 0"/>
      <div style="font-weight:bold;font-size:11px">TECNOLOGÍA DE LA INFORMACIÓN</div>
      <div style="font-size:10px;color:#555">FIRMA QUIEN ENTREGA</div>
    </td>
  </tr>
</table>

</body>
</html>"""
def _enviar_correo_acta(destinatario, nombre_colaborador, pdf_bytes, nombre_archivo):
    """
    Envía el correo con el PDF adjunto.
    Se ejecuta en un hilo separado para no bloquear la respuesta HTTP.
    """
    remitente = settings.EMAIL_HOST_USER
    password  = settings.EMAIL_HOST_PASSWORD
    host      = settings.EMAIL_HOST
    port      = settings.EMAIL_PORT

    # ── Armar el mensaje ──────────────────────────────────
    msg = MIMEMultipart('mixed')
    msg['Subject'] = 'ACTA DE DISPOSITIVOS TECNOLÓGICOS'
    msg['From']    = remitente
    msg['To']      = destinatario

    # ── Cuerpo HTML del correo ────────────────────────────
    cuerpo_html = f"""
    <html><body style="font-family:Arial,sans-serif;font-size:14px;color:#222;line-height:1.6">

      <div style="max-width:600px;margin:0 auto;border:1px solid #e5e7eb;border-radius:8px;overflow:hidden">

        <!-- Cabecera azul -->
        <div style="background:#1e3a5f;padding:24px 30px">
          <h2 style="margin:0;color:#fff;font-size:18px">
            📄 Acta de Dispositivos Tecnológicos
          </h2>
        </div>

        <!-- Cuerpo -->
        <div style="padding:30px">
          <p>Cordial saludo,</p>
          <p>
            Adjunto encontrará el acta de entrega/devolución de equipos tecnológicos
            correspondiente al colaborador <strong>{nombre_colaborador}</strong>,
            la cual hace constancia de la gestión realizada por el área de TI.
          </p>

          <div style="background:#f0f4ff;border-left:4px solid #1e3a5f;padding:14px 18px;margin:20px 0;border-radius:0 6px 6px 0">
            <strong>📎 Documento adjunto:</strong> {nombre_archivo}
          </div>

          <p>Por favor revise el documento y consérvelo para sus registros.</p>
          <p>Gracias,<br><strong>Área de Tecnología e Informática</strong></p>
        </div>

        <!-- Pie de página -->
        <div style="background:#f8fafc;border-top:1px solid #e5e7eb;padding:16px 30px">
          <p style="margin:0;font-size:11px;color:#e53e3e">
            <strong>⚠️ Por favor, no responda a este mensaje, ha sido enviado de forma automática.
            Si desea ponerse en contacto con nosotros para comentarnos alguna incidencia o mejora
            de este servicio, por favor escríbanos a
            <a href="mailto:dirsistemas@montacargasamym.com" style="color:#e53e3e">
            dirsistemas@montacargasamym.com</a></strong>
          </p>
        </div>

      </div>
    </body></html>
    """

    # Parte alternativa para el cuerpo HTML
    parte_alternativa = MIMEMultipart('alternative')
    parte_alternativa.attach(MIMEText(cuerpo_html, 'html', 'utf-8'))
    msg.attach(parte_alternativa)

    # ── Adjuntar el PDF ───────────────────────────────────
    adjunto = MIMEBase('application', 'pdf')
    adjunto.set_payload(pdf_bytes)
    encoders.encode_base64(adjunto)
    adjunto.add_header('Content-Disposition', f'attachment; filename="{nombre_archivo}"')
    adjunto.add_header('Content-Type', 'application/pdf', name=nombre_archivo)
    msg.attach(adjunto)

    # ── Enviar via SMTP (Outlook/Microsoft 365) ───────────
    try:
        with smtplib.SMTP(host, port, timeout=15) as servidor:
            servidor.ehlo()
            servidor.starttls()          # cifrado TLS obligatorio en Outlook
            servidor.ehlo()
            servidor.login(remitente, password)
            servidor.sendmail(remitente, destinatario, msg.as_bytes())
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'[CORREO ACTA] Error enviando a {destinatario}: {e}')   
    
    
    
# ── Correos disparados por Signals.py (post_save de Requerimiento) ────────
# IMPORTANTE: se ejecutan en un hilo aparte (igual que _enviar_correo_acta).
# Así, si el SMTP falla, el error queda aislado en el hilo y NUNCA rompe el
# r.save(using='requerimientos') que dispara la señal — eso es lo que antes
# tumbaba la respuesta HTTP de api_req_tic_accion y dejaba los indicadores
# desactualizados.

def _enviar_correo_solucion(req):
    """
    Notifica al solicitante que su requerimiento fue solucionado (IdEstado -> 4).
    Llamada por Signals.py::_notificar_solucion.
    """
    if not req.Email:
        logging.getLogger(__name__).warning(
            f'[CORREO SOLUCION] Requerimiento {req.codigo()} no tiene Email, no se envía.'
        )
        return

    hilo = threading.Thread(
        target=_enviar_correo_solucion_smtp,
        args=(req.Email, req.NombreUsuario or '', req.codigo(), req.Solucion or ''),
        daemon=True,
    )
    hilo.start()


def _enviar_correo_solucion_smtp(destinatario, nombre_usuario, codigo, solucion):
    remitente = settings.EMAIL_HOST_USER
    password  = settings.EMAIL_HOST_PASSWORD
    host      = settings.EMAIL_HOST
    port      = settings.EMAIL_PORT

    msg = MIMEMultipart('mixed')
    msg['Subject'] = f'Requerimiento solucionado — {codigo}'
    msg['From']    = remitente
    msg['To']      = destinatario

    cuerpo_html = f"""
    <html><body style="font-family:Arial,sans-serif;font-size:14px;color:#222;line-height:1.6">
      <div style="max-width:600px;margin:0 auto;border:1px solid #e5e7eb;border-radius:8px;overflow:hidden">
        <div style="background:#355EAB;padding:24px 30px">
          <h2 style="margin:0;color:#fff;font-size:18px">Requerimiento solucionado</h2>
        </div>
        <div style="padding:30px">
          <p>Hola <strong>{nombre_usuario}</strong>,</p>
          <p>Ya se le dio solución a tu requerimiento <strong>{codigo}</strong>.</p>
          <div style="background:#f0f4ff;border-left:4px solid #355EAB;padding:14px 18px;margin:20px 0;border-radius:0 6px 6px 0">
            <strong>Solución:</strong> {solucion or '(sin detalle)'}
          </div>
          <p>Gracias,<br><strong>Área de Tecnología e Informática</strong></p>
        </div>
      </div>
    </body></html>
    """
    parte_alternativa = MIMEMultipart('alternative')
    parte_alternativa.attach(MIMEText(cuerpo_html, 'html', 'utf-8'))
    msg.attach(parte_alternativa)

    try:
        with smtplib.SMTP(host, port, timeout=15) as servidor:
            servidor.ehlo()
            servidor.starttls()
            servidor.ehlo()
            servidor.login(remitente, password)
            servidor.sendmail(remitente, destinatario, msg.as_bytes())
    except Exception as e:
        logging.getLogger(__name__).error(f'[CORREO SOLUCION] Error enviando a {destinatario}: {e}')


def _enviar_correo_asignacion(req, es_reasignacion=False):
    """
    Notifica al técnico que se le asignó (o reasignó) un requerimiento.
    Llamada por Signals.py::_notificar_asignacion.
    """
    if not req.IdUsuarioAsig:
        return

    try:
        tecnico = Usuario.objects.using('requerimientos').get(IdUsuario=req.IdUsuarioAsig)
    except Usuario.DoesNotExist:
        logging.getLogger(__name__).warning(
            f'[CORREO ASIGNACION] Usuario IdUsuario={req.IdUsuarioAsig} no existe (Requerimiento {req.codigo()}).'
        )
        return

    if not tecnico.Email:
        logging.getLogger(__name__).warning(
            f'[CORREO ASIGNACION] Técnico {tecnico.NombreCompleto} (IdUsuario={tecnico.IdUsuario}) no tiene Email.'
        )
        return

    hilo = threading.Thread(
        target=_enviar_correo_asignacion_smtp,
        args=(tecnico.Email, tecnico.NombreCompleto or '', req.codigo(),
              req.NombreUsuario or '', req.Requerimiento or '', es_reasignacion),
        daemon=True,
    )
    hilo.start()


def _enviar_correo_asignacion_smtp(destinatario, nombre_tecnico, codigo, solicitante, descripcion, es_reasignacion):
    remitente = settings.EMAIL_HOST_USER
    password  = settings.EMAIL_HOST_PASSWORD
    host      = settings.EMAIL_HOST
    port      = settings.EMAIL_PORT

    titulo = 'Requerimiento reasignado' if es_reasignacion else 'Nuevo requerimiento asignado'

    msg = MIMEMultipart('mixed')
    msg['Subject'] = f'{titulo} — {codigo}'
    msg['From']    = remitente
    msg['To']      = destinatario

    cuerpo_html = f"""
    <html><body style="font-family:Arial,sans-serif;font-size:14px;color:#222;line-height:1.6">
      <div style="max-width:600px;margin:0 auto;border:1px solid #e5e7eb;border-radius:8px;overflow:hidden">
        <div style="background:#355EAB;padding:24px 30px">
          <h2 style="margin:0;color:#fff;font-size:18px">{titulo}</h2>
        </div>
        <div style="padding:30px">
          <p>Hola <strong>{nombre_tecnico}</strong>,</p>
          <p>Se te {'reasignó' if es_reasignacion else 'asignó'} el requerimiento <strong>{codigo}</strong>.</p>
          <div style="background:#f0f4ff;border-left:4px solid #355EAB;padding:14px 18px;margin:20px 0;border-radius:0 6px 6px 0">
            <strong>Solicitante:</strong> {solicitante}<br>
            <strong>Descripción:</strong> {descripcion}
          </div>
          <p>Por favor revísalo y dale seguimiento.</p>
          <p>Gracias,<br><strong>Área de Tecnología e Informática</strong></p>
        </div>
      </div>
    </body></html>
    """
    parte_alternativa = MIMEMultipart('alternative')
    parte_alternativa.attach(MIMEText(cuerpo_html, 'html', 'utf-8'))
    msg.attach(parte_alternativa)

    try:
        with smtplib.SMTP(host, port, timeout=15) as servidor:
            servidor.ehlo()
            servidor.starttls()
            servidor.ehlo()
            servidor.login(remitente, password)
            servidor.sendmail(remitente, destinatario, msg.as_bytes())
    except Exception as e:
        logging.getLogger(__name__).error(f'[CORREO ASIGNACION] Error enviando a {destinatario}: {e}')


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_acta_guardar(request, colaborador_id):
    """Guarda el acta y envía el PDF por correo automáticamente."""
    c = get_object_or_404(Colaborador, pk=colaborador_id)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')

    required = ['tipo', 'proceso', 'correo']
    for f in required:
        if not body.get(f):
            return _json_err(f'Campo requerido: {f}')

    # Selección parcial de dispositivos (colaborador con varios asignados y
    # solo se está devolviendo/incluyendo algunos). Si no llega, se usa el
    # comportamiento de siempre: todos los dispositivos asignados al colaborador.
    dispositivos_ids = body.get('dispositivos_ids')
    if dispositivos_ids is not None:
        try:
            dispositivos_ids = [int(x) for x in dispositivos_ids]
        except (TypeError, ValueError):
            return _json_err('dispositivos_ids inválido')

    # 1. Guardar el acta en BD (igual que antes)
    acta = Acta.objects.create(
        g217_colaborador  = c,
        g217_tipo         = body['tipo'],
        g217_proceso      = body['proceso'],
        g217_correo       = body['correo'],
        g217_firma_recibe  = body.get('firma_recibe', ''),
        g217_firma_entrega = body.get('firma_entrega', ''),
    )
    c.g215_correo = body['correo']
    c.save(update_fields=['g215_correo'])

    # 2. Construir los dispositivos (reutilizando la misma lógica de api_acta_detalle)
    #    Si llegó dispositivos_ids, el acta solo cubre esos (los demás asignados
    #    al colaborador quedan intactos, sin aparecer en el acta ni liberarse).
    asignaciones_acta = AsignacionColaborador.objects.filter(g216_colaborador=c)
    if dispositivos_ids:
        asignaciones_acta = asignaciones_acta.filter(g216_dispositivo_id__in=dispositivos_ids)

    dispositivos = []
    for asig in asignaciones_acta.select_related(
        'g216_dispositivo__g212_tipo',
        'g216_dispositivo__caract_pc__g222_procesador',
        'g216_dispositivo__caract_pc__g222_so',
        'g216_dispositivo__caract_pc__g222_antivirus',
        'g216_dispositivo__caract_pc__g222_licencia',
        'g216_dispositivo__caract_pc__g222_tipo_disco',
        'g216_dispositivo__caract_pc__g222_almacenamiento',
        'g216_dispositivo__caract_movil__g223_operador',
        'g216_dispositivo__caract_movil__g223_almacenamiento',
        'g216_dispositivo__caract_pantalla',
        'g216_dispositivo__caract_impresora__g225_tipo_impresora',
        'g216_dispositivo__caract_periferico',
        'g216_dispositivo__caract_licencia',
        'g216_dispositivo__caract_videobeam',
    ):
        d = asig.g216_dispositivo
        dispositivos.append({
            'tipo':            d.g212_tipo.g200_tipo_dispositivo if d.g212_tipo else '—',
            'serial':          d.g212_serial,
            'nombre':          d.g212_nombre_equipo or '—',
            'caracteristicas': _get_caracteristicas_acta(d),
        })

    # 2a. Snapshot: guardar exactamente qué dispositivos quedaron incluidos en
    #     esta acta, para que "ver acta" los muestre siempre igual sin importar
    #     cambios posteriores en las asignaciones del colaborador.
    ActaDispositivo.objects.bulk_create([
        ActaDispositivo(g234_acta=acta, g234_dispositivo_id=asig.g216_dispositivo_id)
        for asig in asignaciones_acta
    ])

    # 2b. Si el acta es de DEVOLUCIÓN → liberar dispositivos y actualizar colaborador
    #     (mismo patrón que api_asignacion_eliminar)
    ESTADO_DISP_LIBRE     = 1   # HABILITADO  (dispositivo libre)
    ESTADO_COLAB_CON_DISP = 6   # DISPOSITIVOS OTORGADOS
    ESTADO_COLAB_SIN_DISP = 10  # SIN ASIGNACIONES

    if 'DEVOLU' in body['tipo'].upper():
        with transaction.atomic():
            asignaciones = AsignacionColaborador.objects.filter(g216_colaborador=c)
            if dispositivos_ids:
                asignaciones = asignaciones.filter(g216_dispositivo_id__in=dispositivos_ids)
            dispositivos_devueltos = list(asignaciones.values_list('g216_dispositivo_id', flat=True))
            asignaciones.delete()

            for dispositivo_id in dispositivos_devueltos:
                dev = Dispositivo.objects.filter(pk=dispositivo_id).first()
                if not dev:
                    continue
                otras = AsignacionColaborador.objects.filter(g216_dispositivo_id=dispositivo_id).exists()
                if not otras:
                    dev.g212_estado_id = ESTADO_DISP_LIBRE
                    dev.save(update_fields=['g212_estado_id'])
                _registrar_historial_auto(
                    dispositivo    = dev,
                    nombre_novedad = 'DEVOLUCIÓN',
                    responsable    = request.user.get_full_name() or request.user.username,
                    observaciones  = f'Devuelto por: {c.g215_nombre} — Acta #{acta.g217_id}',
                    co             = dev.g212_co,
                )

            total = AsignacionColaborador.objects.filter(g216_colaborador=c).count()
            c.g215_estado_id = ESTADO_COLAB_CON_DISP if total > 0 else ESTADO_COLAB_SIN_DISP
            c.save(update_fields=['g215_estado_id'])

    # 3. Logo
    logo_b64   = ''
    logo_path  = os.path.join(settings.BASE_DIR, 'index', 'static', 'img', 'imagen.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_b64 = 'data:image/png;base64,' + base64.b64encode(f.read()).decode()

    # 4. Generar el HTML → PDF con WeasyPrint
    html_acta  = _construir_html_acta(acta, c, dispositivos, logo_b64)
    buffer = BytesIO()
    pisa.CreatePDF(html_acta, dest=buffer)
    pdf_bytes = buffer.getvalue()

    # 5. Enviar el correo en segundo plano (no bloquea la respuesta)
    nombre_archivo = f"Acta_{acta.g217_tipo}_{c.g215_nombre.replace(' ', '_')}_{acta.g217_id}.pdf"
    hilo = threading.Thread(
        target=_enviar_correo_acta,
        args=(body['correo'], c.g215_nombre, pdf_bytes, nombre_archivo),
        daemon=True,
    )
    hilo.start()

    return _json_ok({'acta_id': acta.g217_id})




def _get_caracteristicas_acta(d):
    """
    Retorna dict de características YA FORMATEADO PARA MOSTRAR (claves en
    MAYÚSCULAS, valores resueltos a texto, sin IDs) — uso exclusivo del PDF
    del Acta (_construir_html_acta) y la vista de Acta guardada.

    NO USAR para el formulario de Editar ni el modal de Detalle: para eso
    existe _get_caracteristicas(d) (más arriba en este archivo), que
    devuelve claves técnicas con sufijo _id que el formulario necesita
    para precargar los <select>. Antes ambas funciones se llamaban igual
    y esta (definida después) sobrescribía silenciosamente a la otra,
    rompiendo el precargado de Editar, el Detalle y la exportación a Excel.
    """
    tipo = d.g212_tipo.g200_tipo_dispositivo.upper() if d.g212_tipo else ''
    chars = {}

    if tipo in ('TORRE DE ESCRITORIO', 'PORTATIL'):
        pc = getattr(d, 'caract_pc', None)
        if pc:
            chars = {
                'NOMBRE':        d.g212_nombre_equipo or '—',
                'PROCESADOR':    pc.g222_procesador.g209_procesador if pc.g222_procesador else '—',
                'SO':            pc.g222_so.g210_so if pc.g222_so else '—',
                'RAM':           f"{pc.g222_ram} GB" if pc.g222_ram else '—',
                'TIPO DISCO':    pc.g222_tipo_disco.g231_tipo_disco if pc.g222_tipo_disco else '—',
                'ALMACENAMIENTO': pc.g222_almacenamiento.g219_almacenamiento if pc.g222_almacenamiento else '—',
                'ANTIVIRUS':     pc.g222_antivirus.g208_antivirus if pc.g222_antivirus else '—',
                'OFFICE':        pc.g222_licencia.g211_office if pc.g222_licencia else '—',
                'CORREO / KEY OFFICE': f"{pc.g222_correo_office or '—'} / {pc.g222_key_office or '—'}",
                'ACTIVO':        pc.g222_activo or '—',
            }
            if tipo == 'PORTATIL':
                chars['PULGADAS'] = str(pc.g222_pulgadas) if pc.g222_pulgadas else '—'

    elif tipo in ('CELULAR', 'TABLET', 'MODEM WIFI', 'SIMCARD', 'TELEFONO FIJO'):
        mov = getattr(d, 'caract_movil', None)
        if mov:
            chars = {
                'NÚMERO':        mov.g223_numero_linea or '—',
                'OPERADOR':      mov.g223_operador.g221_operador if mov.g223_operador else '—',
                'PLAN DE DATOS': mov.g223_plan_datos or '—',
                'IMEI 1':        mov.g223_imei1 or '—',
                'IMEI 2':        mov.g223_imei2 or '—',
                'CUENTA GMAIL':  mov.g223_cuenta_gmail or '—',
                'CONTRASEÑA':    mov.g223_contrasena_gmail or '—',
            }
            if tipo in ('MODEM WIFI',):
                chars['ALMACENAMIENTO'] = mov.g223_almacenamiento.g219_almacenamiento if mov.g223_almacenamiento else '—'

    elif tipo == 'PANTALLA':
        pan = getattr(d, 'caract_pantalla', None)
        if pan:
            chars = {
                'PULGADAS':   str(pan.g224_pulgadas) if pan.g224_pulgadas else '—',
                'RESOLUCIÓN': pan.g224_resolucion or '—',
            }

    elif tipo == 'IMPRESORA':
        imp = getattr(d, 'caract_impresora', None)
        if imp:
            chars = {
                'TIPO':    imp.g225_tipo_impresora.g229_tipo_impresora if imp.g225_tipo_impresora else '—',
                'FUNCIÓN': imp.g225_funcion or '—',
            }

    elif tipo == 'PERIFERICO':
        per = getattr(d, 'caract_periferico', None)
        if per:
            chars = {
                'BASE':        'SÍ' if per.g226_incluye_base else 'NO',
                'TECLADO':     'SÍ' if per.g226_incluye_teclado else 'NO',
                'MOUSE':       'SÍ' if per.g226_incluye_mouse else 'NO',
                'AURICULARES': 'SÍ' if per.g226_incluye_auriculares else 'NO',
                'CARGADOR':    'SÍ' if per.g226_incluye_cargador else 'NO',
                'DESCRIPCIÓN': per.g226_descripcion_adicional or '—',
            }

    elif tipo == 'LICENCIA OFFICE':
        lic = getattr(d, 'caract_licencia', None)
        if lic:
            chars = {
                'SOFTWARE':    lic.g227_software or '—',
                'VERSIÓN':     lic.g227_version or '—',
                'KEY':         lic.g227_key or '—',
                'CORREO':      lic.g227_correo or '—',
                'VENCIMIENTO': lic.g227_fecha_vencimiento.strftime('%d/%m/%Y') if lic.g227_fecha_vencimiento else '—',
            }

    # Campos comunes a todos
    chars['VALOR PROMEDIO']       = f"$ {d.g212_valor_promedio:,.0f}" if d.g212_valor_promedio else '—'
    chars['VALOR ARRENDAMIENTO']  = f"$ {d.g212_valor_arrendamiento:,.0f}" if d.g212_valor_arrendamiento else '—'
    chars['MARCA']                = d.g212_marca.g202_marca if d.g212_marca else '—'

    return chars
# FUNSION API DE ACTA DE DETALLE 
@login_required(login_url='login')
@require_http_methods(['GET'])
def api_acta_detalle(request, acta_id):
    acta = get_object_or_404(Acta, pk=acta_id)
    colaborador = acta.g217_colaborador

    # Dispositivos que realmente quedaron incluidos en ESTA acta:
    #  1. Snapshot (ActaDispositivo) — actas creadas después de este fix.
    #  2. Actas viejas de DEVOLUCIÓN sin snapshot → se reconstruye desde el
    #     historial de auditoría, que ya quedó marcado con el número de acta.
    #  3. Cualquier otra acta vieja (Entrega/otro tipo, sin snapshot) → se
    #     mantiene el comportamiento anterior (asignaciones vigentes del
    #     colaborador), porque no hay forma de recuperar ese dato retroactivamente.
    disp_ids = list(
        ActaDispositivo.objects.filter(g234_acta=acta).values_list('g234_dispositivo_id', flat=True)
    )
    if not disp_ids:
        if 'DEVOLU' in (acta.g217_tipo or '').upper():
            disp_ids = list(
                HistorialEquipo.objects.filter(g214_observaciones__icontains=f'Acta #{acta.g217_id}')
                .values_list('g214_dispositivo_id', flat=True)
            )
        else:
            disp_ids = list(
                AsignacionColaborador.objects.filter(g216_colaborador=colaborador)
                .values_list('g216_dispositivo_id', flat=True)
            )

    dispositivos = []
    for d in Dispositivo.objects.filter(pk__in=disp_ids).select_related(
        'g212_tipo',
        'g212_marca',
        'caract_pc__g222_procesador',
        'caract_pc__g222_so',
        'caract_pc__g222_antivirus',
        'caract_pc__g222_licencia',
        'caract_pc__g222_tipo_disco',
        'caract_pc__g222_almacenamiento',
        'caract_movil__g223_operador',
        'caract_movil__g223_almacenamiento',
        'caract_pantalla',
        'caract_impresora__g225_tipo_impresora',
        'caract_periferico',
        'caract_licencia',
        'caract_videobeam',
    ):
        dispositivos.append({
            'tipo':            d.g212_tipo.g200_tipo_dispositivo if d.g212_tipo else '—',
            'serial':          d.g212_serial,
            'nombre':          d.g212_nombre_equipo or '—',
            'caracteristicas': _get_caracteristicas_acta(d),
        })

    #  Logo
    logo_b64 = ''
    logo_path = os.path.join(settings.BASE_DIR, 'index', 'static', 'img', 'imagen.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_b64 = 'data:image/png;base64,' + base64.b64encode(f.read()).decode()
    

    return _json_ok({
        'id':            acta.g217_id,
        'tipo':          acta.g217_tipo,
        'proceso':       acta.g217_proceso,
        'correo':        acta.g217_correo,
        'fecha':         acta.g217_fecha.strftime('%d/%m/%Y %H:%M') if acta.g217_fecha else '—',
        'firma_recibe':  acta.g217_firma_recibe or '',
        'firma_entrega': acta.g217_firma_entrega or '',
        'colaborador': {
            'nombre':    colaborador.g215_nombre,
            'documento': colaborador.g215_documento,
            'cargo':     colaborador.g215_cargo,
            'co':        f"{colaborador.g215_co.g207_co} — {colaborador.g215_co.g207_descripcion_co}" if colaborador.g215_co else '—',
        },
        'dispositivos': dispositivos,
        'logo':          logo_b64,  
    })




#  DASHBOARD — Estadísticas generales y mapa

@login_required(login_url='login')
@require_http_methods(['GET'])
@login_required(login_url='login')
@require_http_methods(['GET'])
def api_exportar_inventario(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    #  Filtros desde la URL (los mismos que el inventario) 
    qs = Dispositivo.objects.select_related(
        'g212_tipo', 'g212_marca', 'g212_propietario',
        'g212_estado', 'g212_co', 'g212_departamento', 'g212_municipio'
    )
    q       = request.GET.get('q', '').strip()
    tipo_id = request.GET.get('tipo', '').strip()
    estado  = request.GET.get('estado', '').strip()

    if q:
        qs = qs.filter(
            Q(g212_serial__icontains=q) |
            Q(g212_propietario__g203_propietario__icontains=q) |
            Q(g212_marca__g202_marca__icontains=q)
        )
    if tipo_id:
        qs = qs.filter(g212_tipo_id=tipo_id)
    if estado:
        qs = qs.filter(g212_estado_id=estado)

    qs = qs.order_by('g212_tipo__g200_tipo_dispositivo', 'g212_serial')

    #  Detectar grupo según tipos presentes 
    tipos_presentes = set(
        qs.values_list('g212_tipo__g200_tipo_dispositivo', flat=True).distinct()
    )

    GRUPO_PC        = {'TORRE DE ESCRITORIO', 'PORTATIL'}
    GRUPO_MOVIL     = {'CELULAR', 'TABLET', 'MODEM WIFI', 'SIMCARD', 'TELEFONO FIJO'}
    GRUPO_PANTALLA  = {'PANTALLA'}
    GRUPO_IMPRESORA = {'IMPRESORA'}
    GRUPO_PERIFERICO= {'PERIFERICO'}
    GRUPO_LICENCIA  = {'LICENCIA OFFICE'}

    def _detectar_grupo():
        for t in tipos_presentes:
            if t in GRUPO_PC:         return 'pc'
            if t in GRUPO_MOVIL:      return 'movil'
            if t in GRUPO_PANTALLA:   return 'pantalla'
            if t in GRUPO_IMPRESORA:  return 'impresora'
            if t in GRUPO_PERIFERICO: return 'periferico'
            if t in GRUPO_LICENCIA:   return 'licencia'
        return 'general'

    grupo = _detectar_grupo()

    #  Columnas base + columnas según grupo 
    BASE_HEADERS = [
        'Serial', 'Tipo', 'Marca', 'Propietario', 'Estado',
        'Centro Operaciones', 'Nombre Equipo',
        'Valor Promedio', 'Valor Arrendamiento',
        'Departamento', 'Municipio', 'Observaciones', 'Fecha Registro',
    ]
    EXTRA_HEADERS = {
        'pc':         ['Procesador', 'SO', 'RAM', 'Tipo Disco', 'Almacenamiento',
                       'Antivirus', 'Office', 'Correo Office', 'Activo'],
        'movil':      ['Número Línea', 'Operador', 'IMEI 1', 'IMEI 2',
                       'Plan Datos', 'Cuenta Gmail', 'Contraseña'],
        'pantalla':   ['Pulgadas', 'Resolución'],
        'impresora':  ['Tipo Impresora', 'Función'],
        'periferico': ['Base', 'Teclado', 'Mouse', 'Auriculares', 'Cargador'],
        'licencia':   ['Software', 'Versión', 'Key', 'Correo', 'Vencimiento'],
        'general':    [],
    }
    headers = BASE_HEADERS + EXTRA_HEADERS.get(grupo, [])

    #  Workbook 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    header_fill = PatternFill('solid', fgColor='1B4698')
    header_font = Font(color='FFFFFF', bold=True)
    center      = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = center

    #  Helpers 
    def _res(Model, pk, attr):
        try:
            return getattr(Model.objects.get(pk=pk), attr) if pk else ''
        except Exception:
            return ''

    #  Filas 
    fill_alt = PatternFill('solid', fgColor='EEF2FF')

    for row_num, d in enumerate(qs, 2):
        caract = _get_caracteristicas(d)

        base = [
            d.g212_serial or '',
            d.g212_tipo.g200_tipo_dispositivo if d.g212_tipo else '',
            d.g212_marca.g202_marca if d.g212_marca else '',
            d.g212_propietario.g203_propietario if d.g212_propietario else '',
            d.g212_estado.g201_descripcion if d.g212_estado else '',
            f"{d.g212_co.g207_co} — {d.g212_co.g207_descripcion_co}" if d.g212_co else '',
            d.g212_nombre_equipo or '',
            float(d.g212_valor_promedio)      if d.g212_valor_promedio      else '',
            float(d.g212_valor_arrendamiento) if d.g212_valor_arrendamiento else '',
            d.g212_departamento.g204_departamento if d.g212_departamento else '',
            d.g212_municipio.g205_municipio if d.g212_municipio else '',
            d.g212_observaciones or '',
            d.g212_fecha_registro.strftime('%d/%m/%Y %H:%M') if d.g212_fecha_registro else '',
        ]

        if grupo == 'pc':
            from dashboard.models import (Procesador, SistemaOperativo,
                                          Antivirus, LicenciaOffice,
                                          Opciones, Almacenamiento)
            extra = [
                _res(Procesador,       caract.get('procesador_id'),    'g209_procesador'),
                _res(SistemaOperativo, caract.get('so_id'),            'g210_so'),
                _res(Opciones,         caract.get('ram_id'),           'g218_opciones'),
                _res(Opciones,         caract.get('tipo_disco_id'),    'g218_opciones'),
                _res(Almacenamiento,   caract.get('almacenamiento_id'),'g219_almacenamiento'),
                _res(Antivirus,        caract.get('antivirus_id'),     'g208_antivirus'),
                _res(LicenciaOffice,   caract.get('licencia_id'),      'g211_office'),
                caract.get('correo_office', ''),
                caract.get('activo', ''),
            ]
        elif grupo == 'movil':
            from dashboard.models import Operador
            extra = [
                caract.get('numero_linea', ''),
                _res(Operador, caract.get('operador_id'), 'g221_operador'),
                caract.get('imei1', ''),
                caract.get('imei2', ''),
                caract.get('plan_datos', ''),
                caract.get('cuenta_gmail', ''),
                caract.get('contrasena_gmail', ''),
            ]
        elif grupo == 'pantalla':
            extra = [caract.get('pulgadas', ''), caract.get('resolucion', '')]
        elif grupo == 'impresora':
            from dashboard.models import TipoImpresora
            extra = [
                _res(TipoImpresora, caract.get('tipo_impresora_id'), 'g229_tipo_impresora'),
                caract.get('funcion', ''),
            ]
        elif grupo == 'periferico':
            extra = [
                'SÍ' if caract.get('incluye_base')        else 'NO',
                'SÍ' if caract.get('incluye_teclado')     else 'NO',
                'SÍ' if caract.get('incluye_mouse')       else 'NO',
                'SÍ' if caract.get('incluye_auriculares') else 'NO',
                'SÍ' if caract.get('incluye_cargador')    else 'NO',
            ]
        elif grupo == 'licencia':
            extra = [
                caract.get('software', ''),
                caract.get('version', ''),
                caract.get('key', ''),
                caract.get('correo', ''),
                caract.get('fecha_vencimiento', ''),
            ]
        else:
            extra = []

        fila = base + extra
        for col, val in enumerate(fila, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            if row_num % 2 == 0:
                cell.fill = fill_alt

    #  Ajustar anchos + freeze header 
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    ws.freeze_panes = 'A2'

    #  Respuesta 
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventario_systraker.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_dashboard_stats(request):
    """
    Estadísticas para el panel de equipos del dashboard:
    - Conteo por tipo de dispositivo (para el carrusel)
    - Totales activos / inactivos
    - Ubicaciones geográficas para el mapa
    """
    # Carrusel: todos los tipos activos, con conteo (0 si no tienen dispositivos)
    ICONOS = {
        'TORRE DE ESCRITORIO': 'https://img.icons8.com/3d-fluency/96/desktop.png',
        'PORTATIL':            'https://img.icons8.com/3d-fluency/96/laptop.png',
        'PORTÁTIL':            'https://img.icons8.com/3d-fluency/96/laptop.png',
        'PANTALLA':            'https://img.icons8.com/3d-fluency/96/monitor.png',
        'CELULAR':             'https://img.icons8.com/3d-fluency/96/iphone14-pro.png',
        'MODEM WIFI':          'https://img.icons8.com/3d-fluency/96/router.png',
        'DIADEMA':             'https://img.icons8.com/3d-fluency/96/headset.png',
        'SIMCARD':             'https://img.icons8.com/3d-fluency/96/sim-card.png',
        'VIDEO BEAM':          'https://img.icons8.com/3d-fluency/96/video-projector.png',
        'TABLET':              'https://img.icons8.com/3d-fluency/96/ipad.png',
        'TELEFONO FIJO':       'https://img.icons8.com/3d-fluency/96/telephone.png',
        'TELÉFONO FIJO':       'https://img.icons8.com/3d-fluency/96/telephone.png',
        'IMPRESORA':           'https://img.icons8.com/3d-fluency/96/print.png',
        'PERIFERICO':          'https://img.icons8.com/3d-fluency/96/usb-2.png',
        'PERIFÉRICO':          'https://img.icons8.com/3d-fluency/96/usb-2.png',
        'LICENCIA OFFICE':     'https://img.icons8.com/3d-fluency/96/ms-office.png',
    }
    # Conteo real de dispositivos por tipo
    conteo_dict = {
        row['g212_tipo__g200_tipo_dispositivo']: row['value']
        for row in (
            Dispositivo.objects
            .exclude(g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS)
            .values('g212_tipo__g200_tipo_dispositivo')
            .annotate(value=Count('g212_id'))
        )
    }

    # Todos los tipos activos, con valor 0 si no tienen dispositivos aún
    tipos = [
        {
            'label': t.g200_tipo_dispositivo,
            'value': conteo_dict.get(t.g200_tipo_dispositivo, 0),
            'src':   ICONOS.get(t.g200_tipo_dispositivo, 'https://img.icons8.com/fluency/96/server.png'),
        }
        for t in TipoDispositivo.objects.filter(g200_estado=True).order_by('g200_tipo_dispositivo')
    ]

    # Activos e inactivos totales
    activos_total   = Dispositivo.objects.exclude(g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS).count()
    inactivos_total = Dispositivo.objects.filter(g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS).count()

    # Ubicaciones para el mapa: usa coordenadas de la BD
    ubicaciones_raw = (
        Dispositivo.objects
        .select_related('g212_municipio', 'g212_tipo')
        .exclude(g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS)
        .exclude(g212_municipio=None)
        .exclude(g212_municipio__g205_latitud=None)
        .exclude(g212_municipio__g205_longitud=None)
        .values(
            'g212_municipio__g205_municipio',
            'g212_municipio__g205_latitud',
            'g212_municipio__g205_longitud',
            'g212_tipo__g200_tipo_dispositivo',
        )
        .annotate(cantidad=Count('g212_id'))
    )

    ubicaciones = []
    for row in ubicaciones_raw:
        lat = row['g212_municipio__g205_latitud']
        lng = row['g212_municipio__g205_longitud']
        if lat and lng:
            ubicaciones.append({
                'lat':      float(lat),
                'lng':      float(lng),
                'ciudad':   row['g212_municipio__g205_municipio'] or '',
                'tipo':     row['g212_tipo__g200_tipo_dispositivo'] or '',
                'cantidad': row['cantidad'],
            })

    ciudades_count = len(set(u['ciudad'] for u in ubicaciones))
    return _json_ok({
        'tipos':     tipos,
        'activos':   activos_total,
        'inactivos': inactivos_total,
        'ubicaciones': ubicaciones,
        'ciudades':  ciudades_count,
    })


# ENDPOINT: ELIMINAR UNA ASIGNACIÓN INDIVIDUAL
# DELETE /api/colaboradores/<colaborador_id>/asignar/<dispositivo_id>/

@login_required(login_url='login')
@require_http_methods(['DELETE'])
def api_asignacion_eliminar(request, colaborador_id, dispositivo_id):
    """
    Elimina una sola asignación colaborador↔dispositivo sin tocar las demás.

    Cambios de estado automáticos:
      - Dispositivo removido → estado HABILITADO (id=1)
      - Si el colaborador queda sin dispositivos → estado SIN ASIGNACIONES (id=10)
      - Si aún le quedan dispositivos → estado DISPOSITIVOS OTORGADOS (id=6)
    """
    # ── IDs de estado (tabla j201_estado) ──────────────────────────
    ESTADO_DISP_LIBRE       = 1   # HABILITADO  (dispositivo libre)
    ESTADO_COLAB_CON_DISP   = 6   # DISPOSITIVOS OTORGADOS
    ESTADO_COLAB_SIN_DISP   = 10  # SIN ASIGNACIONES
    # ───────────────────────────────────────────────────────────────

    c = get_object_or_404(Colaborador, pk=colaborador_id)

    with transaction.atomic():
        deleted, _ = AsignacionColaborador.objects.filter(
            g216_colaborador=c,
            g216_dispositivo_id=dispositivo_id,
        ).delete()
        
        # Historial automático — desasignación individual
        dev_rem = Dispositivo.objects.filter(pk=dispositivo_id).first()
        if dev_rem:
            _registrar_historial_auto(
                dispositivo   = dev_rem,
                nombre_novedad = 'DESASIGNACIÓN DE COLABORADOR',
                responsable   = request.user.get_full_name() or request.user.username,
                observaciones = f'Removido de: {c.g215_nombre}',
                co            = dev_rem.g212_co,
            )

        if deleted == 0:
            return _json_err('Asignación no encontrada', status=404)

        # Solo poner HABILITADO si no tiene otras asignaciones activas
        otras = AsignacionColaborador.objects.filter(
            g216_dispositivo_id=dispositivo_id
        ).exists()
        Dispositivo.objects.filter(pk=dispositivo_id).update(
            g212_estado_id=3 if otras else ESTADO_DISP_LIBRE
        )

        # Recalcular estado del colaborador
        total = AsignacionColaborador.objects.filter(g216_colaborador=c).count()
        c.g215_estado_id = ESTADO_COLAB_CON_DISP if total > 0 else ESTADO_COLAB_SIN_DISP
        c.save(update_fields=['g215_estado_id'])

    return _json_ok({'colaborador_id': colaborador_id, 'asignados': total})



# ─────────────────────────────────────────────────────────────
#  REEMPLAZA tu función api_carga_masiva actual por ESTA.
#  No cambia nombres de campos, modelos, ni la lógica de negocio.
#  Solo agrega:
#    1) Blindaje total: SIEMPRE responde JSON, nunca HTML.
#    2) Normalización uniforme (mayúsculas/tildes/espacios) en TODOS
#       los cruces contra catálogos.
#    3) Saneo de números que Excel rompe (IMEI, teléfonos, seriales).
#
#  Requiere (agregar arriba del archivo, junto a los demás imports,
#  si no lo tienes ya):
#
#      import logging
#      import unicodedata
#      logger = logging.getLogger(__name__)
#
#

import logging
import unicodedata

logger = logging.getLogger(__name__)


def _normalizar(texto):
    """
    Normaliza texto para comparaciones robustas contra catálogos:
    - Convierte a string
    - Quita tildes/acentos
    - Colapsa espacios múltiples
    - Pasa a MAYÚSCULAS
    - Quita espacios al inicio/fin
    Ej: '  Portátil  ' -> 'PORTATIL'   |   'móDem wifi' -> 'MODEM WIFI'
    """
    if texto is None:
        return ''
    s = str(texto).strip()
    if not s:
        return ''
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = ' '.join(s.split())
    return s.upper()


def _normalizar_ram(valor):
    """
    Extrae el valor numérico (GB) del campo RAM, descartando cualquier
    letra o unidad que el usuario haya escrito.
    Ej: '16' -> 16   |   '16GB' -> 16   |   '16 gb' -> 16
    """
    if not valor:
        return None
    match = re.search(r'\d+', str(valor))
    return int(match.group()) if match else None


def _texto_excel_seguro(valor):
    """
    Convierte un valor de celda de Excel a texto de forma segura,
    evitando el problema clásico de números largos (IMEI, teléfonos,
    seriales) que Excel guarda como float y que Python convierte
    a algo como '356938035643809.0' o notación científica.
    """
    if valor is None:
        return ''
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor)
    if isinstance(valor, int):
        return str(valor)
    return str(valor).strip()


@csrf_exempt
@login_required(login_url='login')
@require_http_methods(['POST'])
def api_carga_masiva(request):
    """
    Importa dispositivos desde un archivo Excel (.xlsx / .xls).
    Recibe:
        archivo          — archivo Excel
        tipo_dispositivo — nombre exacto del tipo (Ej: CELULAR, PORTATIL…)

    Respuesta JSON (SIEMPRE, incluso ante errores inesperados):
        { ok: true/false, data|error: ... }
    """
   

    # ─────────────────────────────────────────────────────────
    # BLINDAJE GLOBAL: todo lo de abajo va dentro de este try.
    # Cualquier excepción no prevista (BD caída, encoding raro,
    # Excel corrupto, etc.) termina en _json_err, NUNCA en la
    # página de error HTML de Django.
    # ─────────────────────────────────────────────────────────
    try:
        import openpyxl
        from decimal import Decimal, InvalidOperation

        #  1. Validar archivo 
        archivo = request.FILES.get('archivo')
        if not archivo:
            return _json_err('Se requiere el archivo Excel (campo "archivo")')

        nombre = archivo.name.lower()
        if not (nombre.endswith('.xlsx') or nombre.endswith('.xls')):
            return _json_err('Solo se aceptan archivos .xlsx o .xls')

        #  2. Leer tipo_dispositivo del request 
        tipo_nombre = request.POST.get('tipo_dispositivo', '').strip()
        if not tipo_nombre:
            return _json_err('Se requiere el campo "tipo_dispositivo"')

        tipo_nombre_norm = _normalizar(tipo_nombre)

        tipo_obj = None
        for t in TipoDispositivo.objects.filter(g200_estado=True):
            if _normalizar(t.g200_tipo_dispositivo) == tipo_nombre_norm:
                tipo_obj = t
                break

        if not tipo_obj:
            return _json_err(
                f'El tipo de dispositivo "{tipo_nombre}" no existe en el catálogo. '
                f'Verifica que esté registrado en Tipo Dispositivo.'
            )

        #  3. Leer workbook 
        try:
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
            filas = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return _json_err(f'No se pudo leer el archivo: {e}')

        if len(filas) < 2:
            return _json_err('El archivo no tiene datos (solo encabezado o está vacío)')

        # ── 4. Mapear columnas por nombre (case/tilde-insensitive) ─────
        encabezado = [
            _normalizar(c).lower().replace(' ', '_') if c else ''
            for c in filas[0]
        ]

        ALIAS = {
            'serial':              ['serial', 'serial_del_dispositivo'],
            'marca':               ['marca'],
            'propietario':         ['propietario', 'nombre_del_propietario'],
            'centro_operaciones':  ['centro_operaciones', 'co', 'centro_de_operaciones'],
            'departamento':        ['departamento'],
            'municipio':           ['municipio'],
            'observaciones':       ['observaciones', 'obs'],
            'nombre':              ['nombre', 'nombre_del_equipo', 'nombre_equipo'],
            'valor_promedio':      ['valor_promedio', 'valor'],
            'valor_arrendamiento': ['valor_arrendamiento', 'arrend'],
            'procesador':          ['procesador'],
            'ram':                 ['ram'],
            'disco':               ['disco', 'tipo_disco', 'tipo_de_disco'],
            'almacenamiento':      ['almacenamiento', 'capacidad'],
            'so':                  ['so', 'sistema_operativo'],
            'antivirus':           ['antivirus'],
            'licencia_office':     ['licencia_office', 'licencia'],
            'tipo_licencia':       ['tipo_licencia', 'tipo_de_licencia', 'licencia'],
            'correo_office':       ['correo_office', 'correo'],
            'activo':              ['activo'],
            'pulgadas':            ['pulgadas'],
            'numero_linea':        ['numero_linea', 'numero', 'linea'],
            'operador':            ['operador'],
            'imei1':               ['imei1', 'imei_1', 'imei'],
            'imei2':               ['imei2', 'imei_2'],
            'plan_datos':          ['plan_datos', 'plan_de_datos'],
            'cuenta_email':        ['cuenta_email', 'cuenta_de_email', 'email', 'gmail'],
            'contrasena':          ['contrasena', 'contraseña', 'contrasena_gmail', 'password'],
            'resolucion':          ['resolucion', 'resolución'],
            'tipo_impresora':      ['tipo_impresora'],
            'funcion':             ['funcion', 'función'],
            'base':                ['base'],
            'teclado':             ['teclado'],
            'mouse':               ['mouse'],
            'auriculares':         ['auriculares'],
            'cargador_pc':         ['cargador_pc'],
            'cargador_movil':      ['cargador_movil', 'cargador_móvil'],
            'version':             ['version', 'versión', 'version_office'],
            'lumenes':             ['lumenes', 'lúmenes'],
        }

        def _idx(campo):
            for alias in ALIAS.get(campo, [campo]):
                alias_norm = _normalizar(alias).lower().replace(' ', '_')
                if alias_norm in encabezado:
                    return encabezado.index(alias_norm)
            return None

        idx = {campo: _idx(campo) for campo in ALIAS}

        def _val(fila, campo):
            i = idx.get(campo)
            if i is None or i >= len(fila):
                return ''
            return _texto_excel_seguro(fila[i])

        def _bool(fila, campo):
            v = _normalizar(_val(fila, campo))
            return v in ('SI', 'YES', 'TRUE', '1', 'S')

        def _decimal(fila, campo):
            s = _val(fila, campo).replace(',', '.').replace('$', '').replace(' ', '')
            try:
                return Decimal(s) if s else None
            except InvalidOperation:
                return None

        def _entero(fila, campo):
            s = _val(fila, campo).replace(' ', '')
            try:
                return int(float(s)) if s else None
            except ValueError:
                return None

        # ── 5. Validar columna serial obligatoria ──────────────────────
        if idx['serial'] is None:
            return _json_err(
                f'Columna "serial" no encontrada. '
                f'Columnas detectadas: {", ".join(encabezado)}'
            )

        # ── 6. Cargar cachés normalizados (evita N+1 queries) ──────────
        marcas_cache      = {_normalizar(m.g202_marca): m       for m in Marca.objects.all() if m.g202_marca}
        propiet_cache     = {_normalizar(p.g203_propietario): p for p in Propietario.objects.all() if p.g203_propietario}
        co_cache          = {_normalizar(c.g207_co): c          for c in CentroOperaciones.objects.all() if c.g207_co}
        dpto_cache        = {_normalizar(d.g204_departamento): d for d in Departamento.objects.all() if d.g204_departamento}
        muni_cache        = {
            (_normalizar(m.g205_municipio), m.g205_departamento_id): m
            for m in Municipio.objects.select_related('g205_departamento').all()
            if m.g205_municipio
        }
        proce_cache       = {_normalizar(p.g209_procesador): p  for p in Procesador.objects.all() if p.g209_procesador}
        so_cache          = {_normalizar(s.g210_so): s          for s in SistemaOperativo.objects.all() if s.g210_so}
        antivirus_cache   = {_normalizar(a.g208_antivirus): a   for a in Antivirus.objects.all() if a.g208_antivirus}
        licencia_cache    = {_normalizar(l.g211_office): l      for l in LicenciaOffice.objects.all() if l.g211_office}
        disco_cache       = {_normalizar(d.g231_tipo_disco): d  for d in TipoDisco.objects.all() if d.g231_tipo_disco}
        alm_cache         = {_normalizar(a.g219_almacenamiento): a for a in Almacenamiento.objects.all() if a.g219_almacenamiento}
        operador_cache    = {_normalizar(o.g221_operador): o    for o in Operador.objects.all() if o.g221_operador}
        timpres_cache     = {_normalizar(t.g229_tipo_impresora): t for t in TipoImpresora.objects.all() if t.g229_tipo_impresora}

        estado_default = Estado.objects.filter(
            g201_descripcion__iexact='HABILITADO'
        ).first()

        # ── 7. Helpers para resolver FK por nombre (normalizado) ───────
        def _resolve_marca(fila):
            n = _normalizar(_val(fila, 'marca'))
            if not n:
                return None
            obj = marcas_cache.get(n)
            if not obj:
                obj, _c = Marca.objects.get_or_create(
                    g202_marca__iexact=n,
                    defaults={'g202_marca': n.title(), 'g202_estado': True}
                )
                marcas_cache[n] = obj
            return obj

        def _resolve_dpto_muni(fila):
            dpto_n = _normalizar(_val(fila, 'departamento'))
            muni_n = _normalizar(_val(fila, 'municipio'))
            dpto_obj = dpto_cache.get(dpto_n) if dpto_n else None
            muni_obj = None
            if muni_n and dpto_obj:
                muni_obj = muni_cache.get((muni_n, dpto_obj.g204_id))
            return dpto_obj, muni_obj

        # ── 8. Agrupar tipos por familia de característica ─────────────
        FAMILIA_MOVIL      = {'CELULAR', 'TABLET', 'MODEM WIFI', 'SIMCARD', 'TELEFONO FIJO'}
        FAMILIA_PC         = {'PORTATIL', 'TORRE DE ESCRITORIO', 'TORRE'}
        FAMILIA_PANTALLA   = {'PANTALLA', 'MONITOR'}
        FAMILIA_IMPRESORA  = {'IMPRESORA'}
        FAMILIA_PERIFERICO = {'PERIFERICO'}
        FAMILIA_LICENCIA   = {'LICENCIA OFFICE', 'LICENCIA'}
        FAMILIA_VIDEOBEAM  = {'VIDEO BEAM'}

        tipo_upper = tipo_nombre_norm

        # ── 9. Procesar filas ──────────────────────────────────────────
        creados  = 0
        omitidos = 0
        errores  = []

        for num_fila, fila in enumerate(filas[1:], start=2):

            serial = _normalizar_serial(_val(fila, 'serial'))
            if not serial:
                omitidos += 1
                errores.append({
                    'fila': num_fila,
                    'serial': '(vacío)',
                    'error': 'Fila omitida — campo serial vacío'
                })
                continue

            if Dispositivo.objects.filter(g212_serial=serial).exists():
                omitidos += 1
                errores.append({
                    'fila': num_fila,
                    'serial': serial,
                    'error': 'Serial duplicado — ya existe en inventario'
                })
                continue

            try:
                with transaction.atomic():

                    dpto_obj, muni_obj = _resolve_dpto_muni(fila)
                    co_nombre_n = _normalizar(_val(fila, 'centro_operaciones'))
                    co_obj = co_cache.get(co_nombre_n) if co_nombre_n else None

                    disp = Dispositivo.objects.create(
                        g212_serial          = serial,
                        g212_tipo            = tipo_obj,
                        g212_marca           = _resolve_marca(fila),
                        g212_propietario     = propiet_cache.get(_normalizar(_val(fila, 'propietario'))),
                        g212_estado          = estado_default,
                        g212_co              = co_obj,
                        g212_nombre_equipo   = _val(fila, 'nombre') or None,
                        g212_departamento    = dpto_obj,
                        g212_municipio       = muni_obj,
                        g212_valor_promedio      = _decimal(fila, 'valor_promedio'),
                        g212_valor_arrendamiento = _decimal(fila, 'valor_arrendamiento'),
                        g212_observaciones   = _val(fila, 'observaciones') or None,
                    )

                    if tipo_upper in FAMILIA_MOVIL:
                        op_n = _normalizar(_val(fila, 'operador'))
                        CaracteristicaMovil.objects.create(
                            g223_dispositivo      = disp,
                            g223_numero_linea     = _val(fila, 'numero_linea') or None,
                            g223_operador         = operador_cache.get(op_n) if op_n else None,
                            g223_plan_datos       = _val(fila, 'plan_datos') or None,
                            g223_imei1            = _val(fila, 'imei1') or None,
                            g223_imei2            = _val(fila, 'imei2') or None,
                            g223_cuenta_gmail     = _val(fila, 'cuenta_email') or None,
                            g223_contrasena_gmail = _val(fila, 'contrasena') or None,
                            g223_pulgadas         = _decimal(fila, 'pulgadas'),
                            g223_almacenamiento   = alm_cache.get(_normalizar(_val(fila, 'almacenamiento'))) if _val(fila, 'almacenamiento') else None,
                        )

                    elif tipo_upper in FAMILIA_PC:
                        pro_n = _normalizar(_val(fila, 'procesador'))
                        so_n  = _normalizar(_val(fila, 'so'))
                        ant_n = _normalizar(_val(fila, 'antivirus'))
                        lic_n = _normalizar(_val(fila, 'licencia_office'))
                        dis_n = _normalizar(_val(fila, 'disco'))
                        alm_n = _normalizar(_val(fila, 'almacenamiento'))
                        CaracteristicaPC.objects.create(
                            g222_dispositivo    = disp,
                            g222_procesador     = proce_cache.get(pro_n) if pro_n else None,
                            g222_so             = so_cache.get(so_n) if so_n else None,
                            g222_antivirus      = antivirus_cache.get(ant_n) if ant_n else None,
                            g222_licencia       = licencia_cache.get(lic_n) if lic_n else None,
                            g222_ram            = _normalizar_ram(_val(fila, 'ram')),
                            g222_tipo_disco     = disco_cache.get(dis_n) if dis_n else None,
                            g222_almacenamiento = alm_cache.get(alm_n) if alm_n else None,
                            g222_correo_office  = _val(fila, 'correo_office') or None,
                            g222_activo         = _val(fila, 'activo') or None,
                            g222_pulgadas       = _decimal(fila, 'pulgadas'),
                        )

                    elif tipo_upper in FAMILIA_PANTALLA:
                        CaracteristicaPantalla.objects.create(
                            g224_dispositivo = disp,
                            g224_pulgadas    = _decimal(fila, 'pulgadas'),
                            g224_resolucion  = _val(fila, 'resolucion') or None,
                        )

                    elif tipo_upper in FAMILIA_IMPRESORA:
                        ti_n = _normalizar(_val(fila, 'tipo_impresora'))
                        CaracteristicaImpresora.objects.create(
                            g225_dispositivo    = disp,
                            g225_tipo_impresora = timpres_cache.get(ti_n) if ti_n else None,
                            g225_funcion        = _val(fila, 'funcion') or None,
                        )

                    elif tipo_upper in FAMILIA_PERIFERICO:
                        CaracteristicaPeriferico.objects.create(
                            g226_dispositivo           = disp,
                            g226_incluye_base          = _bool(fila, 'base'),
                            g226_incluye_teclado       = _bool(fila, 'teclado'),
                            g226_incluye_mouse         = _bool(fila, 'mouse'),
                            g226_incluye_auriculares   = _bool(fila, 'auriculares'),
                            g226_incluye_cargador      = _bool(fila, 'cargador_pc') or _bool(fila, 'cargador_movil'),
                            g226_descripcion_adicional = _val(fila, 'observaciones') or None,
                        )

                    elif tipo_upper in FAMILIA_LICENCIA:
                        alm_n_lic = _normalizar(_val(fila, 'almacenamiento'))
                        CaracteristicaLicencia.objects.create(
                            g227_dispositivo = disp,
                            g227_software    = _val(fila, 'tipo_licencia') or None,
                            g227_version     = _val(fila, 'version') or None,
                            g227_key         = None,
                            g227_correo      = None,
                            g227_fecha_vencimiento = None,
                            g227_almacenamiento = alm_cache.get(alm_n_lic) if alm_n_lic else None,
                        )

                    elif tipo_upper in FAMILIA_VIDEOBEAM:
                        CaracteristicasVideoBeam.objects.create(
                            g232_dispositivo = disp,
                            g232_lumenes      = _entero(fila, 'lumenes'),
                        )
                    # Otros tipos sin tabla de características propia: solo j212

                creados += 1

                _registrar_historial_auto(
                    dispositivo    = disp,
                    nombre_novedad = 'INGRESO AL INVENTARIO',
                    responsable    = request.user.get_full_name() or request.user.username,
                    observaciones  = f'Carga masiva — tipo: {tipo_nombre}',
                    co             = disp.g212_co,
                )

            except Exception as e:
                # Error puntual de esta fila: se registra y se sigue con las demás,
                # NO se interrumpe toda la carga.
                omitidos += 1
                errores.append({'fila': num_fila, 'serial': serial, 'error': str(e)})
                logger.warning(f'[CARGA MASIVA] Fila {num_fila} ({serial}) falló: {e}')

        return _json_ok({
            'creados':  creados,
            'omitidos': omitidos,
            'errores':  errores,
        })

    except Exception as e:
        # ── Blindaje final: cualquier error no previsto arriba
        # (BD, permisos, archivo corrupto, encoding, etc.) se
        # captura aquí y SIEMPRE se responde JSON, nunca HTML.
        logger.exception('[CARGA MASIVA] Error inesperado procesando el archivo')
        return _json_err(
            f'Error interno procesando el archivo: {e}',
            status=500
        )
#GESTION DE USUARIOS   
import datetime
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password


DB = 'requerimientos'

@login_required
@require_GET
def api_req_centros_operacion(request):
    """Lista de CentroOperacion para el select del modal de usuarios."""
    cos = CentroOperacion.objects.using(DB).order_by('Descripcion')
    data = [{'id': c.IdCo, 'nombre': c.Descripcion} for c in cos]
    return JsonResponse({'ok': True, 'results': data})


@login_required
@require_GET
def api_req_cargos(request):
    """Lista de Cargo para el select del modal de usuarios."""
    cargos = Cargo.objects.using(DB).order_by('Descripcion')
    data = [{'id': c.IdCargo, 'nombre': c.Descripcion} for c in cargos]
    return JsonResponse({'ok': True, 'results': data})


@login_required
@require_GET
def api_usuarios_req(request):
    search = request.GET.get('q', '').strip()
    page   = int(request.GET.get('page', 1))
    size   = int(request.GET.get('size', 10))

    qs = Usuario.objects.using(DB).filter(Estado=1).order_by('NombreCompleto')
    if search:
        qs = qs.filter(NombreCompleto__icontains=search) | \
             Usuario.objects.using(DB).filter(Estado=1, Cedula__icontains=search)

    total    = qs.count()
    usuarios = list(qs[(page-1)*size : page*size])

    cargo_ids  = [u.IdCargo     for u in usuarios if u.IdCargo]
    co_ids     = [u.IdCO        for u in usuarios if u.IdCO]
    tipo_ids   = [u.TipoUsuario for u in usuarios if u.TipoUsuario]
    cargos_map = {c.IdCargo: c.Descripcion        for c in Cargo.objects.using(DB).filter(IdCargo__in=cargo_ids)}
    cos_map    = {c.IdCo: c.Descripcion           for c in CentroOperacion.objects.using(DB).filter(IdCo__in=co_ids)}
    tipos_map  = {t.idTipoUsuario: t.Descripcion for t in TipoUsuario.objects.using(DB).filter(idTipoUsuario__in=tipo_ids)}

    data = [{
    'id':              u.IdUsuario,
    'cedula':          u.Cedula,
    'nombre':          u.NombreCompleto,
    'cargo':           cargos_map.get(u.IdCargo, ''),
    'cargo_id':        u.IdCargo or '',
    'co':              cos_map.get(u.IdCO, ''),
    'co_id':           u.IdCO or '',
    'correo':          u.Email or '',
    'fecha':           u.FechaCreacion.strftime('%Y-%m-%d') if u.FechaCreacion else '',
    'tipo_usuario':    tipos_map.get(u.TipoUsuario, '') if u.TipoUsuario else '',
    'tipo_usuario_id': u.TipoUsuario or '',
    } for u in usuarios]
    return JsonResponse({'ok': True, 'total': total, 'page': page, 'size': size, 'results': data})

@login_required
@csrf_exempt
@require_POST
def api_usuario_req_crear(request):
    try:
        body     = json.loads(request.body)
        cedula   = body.get('cedula', '').strip()
        nombre   = body.get('nombre', '').strip()
        correo   = body.get('correo', '').strip()
        password = body.get('password', '').strip()
        tipo = body.get('tipo_usuario') or None 
        id_co    = body.get('co_id', '').strip()        # ← era 'id_co'
        id_cargo = body.get('cargo_id') or None          # ← era 'id_cargo'

        if not cedula or not nombre or not password:
            return JsonResponse({'ok': False, 'error': 'Cédula, nombre y contraseña son requeridos'}, status=400)
        if not id_co:
            return JsonResponse({'ok': False, 'error': 'El Centro de Operación es requerido'}, status=400)
        if Usuario.objects.using(DB).filter(Cedula=cedula).exists():
            return JsonResponse({'ok': False, 'error': 'Ya existe un usuario con esa cédula'}, status=400)

        u = Usuario(
            Cedula         = cedula,
            NombreCompleto = nombre,
            Email          = correo,
            Contrasena     = make_password(password),
            TipoUsuario    = tipo,
            FechaCreacion  = datetime.datetime.now(),
            Estado         = 1,
            IdCO           = id_co,     # ← NUEVO
            IdCargo        = id_cargo,  # ← NUEVO
        )
        u.save(using=DB)
        return JsonResponse({'ok': True, 'id': u.IdUsuario})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)
    
    
@login_required
@require_GET
def api_req_tipos_usuario(request):
    tipos = TipoUsuario.objects.using(DB).order_by('Descripcion')
    data = [{'id': t.idTipoUsuario, 'nombre': t.Descripcion} for t in tipos]
    return JsonResponse({'ok': True, 'results': data})




@login_required
@csrf_exempt
@require_POST
def api_usuario_req_eliminar(request, pk):
    try:
        u        = Usuario.objects.using(DB).get(IdUsuario=pk)
        u.Estado = 0   # baja lógica
        u.save(using=DB)
        return JsonResponse({'ok': True})
    except Usuario.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)
    
    
    
# ─────────────────────────────────────────────
# REQUERIMIENTOS ASIGNADOS AL TÉCNICO EN SESIÓN
# ─────────────────────────────────────────────
from requerimientos.models import Requerimiento, Categoria, SubCategoria,Prioridad

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_mis_req_tic(request):
    """Requerimientos asignados al técnico que inició sesión."""
    from requerimientos.models import (
        Cargo as CargoReq, TipoRequerimiento, Clasificacion as ClasificacionReq,
        ImagenAdjunta, CentroOperacion as CentroOperacionReq,
    )

    user_id = request.session.get('req_user_id')
    if not user_id:
        return _json_err('Tu usuario no está vinculado a la BD de requerimientos', 403)

    ESTADOS       = {1:'Abierto', 2:'Asignado', 3:'En Proceso', 4:'Cerrado', 5:'Eliminado', 6:'Calificado',
                     7:'Pendiente Aprobación', 8:'Rechazado', 9:'Requiere corrección'}
    PRIORIDADES   = {p.IdPrioridad: p.Descripcion for p in Prioridad.objects.using('requerimientos').all()}
    CATEGORIAS    = {c.IdCategoria: c.Descripcion for c in Categoria.objects.using('requerimientos').all()}
    SUBCATEGORIAS = {s.IdSubCategoria: s.Descripcion for s in SubCategoria.objects.using('requerimientos').all()}
    TIPOS         = {t.IdTipoReque: t.Descripcion for t in TipoRequerimiento.objects.using('requerimientos').all()}
    CARGOS        = {c.IdCargo: c.Descripcion for c in CargoReq.objects.using('requerimientos').all()}
    CLASIFICAC    = {c.IdClasificacion: c.Clasificacion for c in ClasificacionReq.objects.using('requerimientos').all()}
    # CO guarda el código corto (ej. "AM1"), se resuelve al nombre completo.
    CENTROS       = {c.IdCo: c.Descripcion for c in CentroOperacionReq.objects.using('requerimientos').all()}

    qs = (Requerimiento.objects
          .using('requerimientos')
          .filter(IdUsuarioAsig=user_id)
          .exclude(IdEstado=5)
          .order_by('-Fecha'))

    # Adjuntos: mismo criterio que api_todos_req_tic — solo se resuelven los
    # que subió este proyecto (ver ADJUNTO_CARPETA en requerimientos/views.py).
    codigos_pagina = [r.Codigo for r in qs]
    ADJUNTOS = {}
    for img in (ImagenAdjunta.objects.using('requerimientos')
                .filter(CodReq__in=codigos_pagina).order_by('IdImagen')):
        nombre_disco = f'{img.IdImagen}_{img.NombreImagen}'
        ADJUNTOS[img.CodReq] = {
            'nombre': img.NombreImagen,
            'url':    f'{settings.MEDIA_URL}requerimientos_adjuntos/{nombre_disco}',
        }

    data = []
    for r in qs:
        adjunto = ADJUNTOS.get(r.Codigo)
        centro_nombre = CENTROS.get(r.CO, r.CO or '—')
        data.append({
            'codigo':             r.codigo(),
            'id':                 r.Codigo,
            'solicitante':        r.NombreUsuario or '—',
            'cedula':             str(r.CedulaUsuario or ''),
            'documento':          r.CedulaUsuario or '',
            'correo':             r.Email or '',
            'cargo':              CARGOS.get(r.Cargo, '—'),
            'co':                 centro_nombre,
            'centro_operacion':   centro_nombre,
            'tiene_adjunto':      bool(adjunto),
            'nombre_adjunto':     adjunto['nombre'] if adjunto else '',
            'url_adjunto':        adjunto['url'] if adjunto else '',
            'requerimiento':      r.Requerimiento or '',
            'tipo_requerimiento': TIPOS.get(r.IdTipoReq, '—'),
            'categoria':          CATEGORIAS.get(r.IdCategoria, '—'),
            'subcategoria':       SUBCATEGORIAS.get(r.IdSubCategoria, '—'),
            'clasificacion':      CLASIFICAC.get(r.Clasificacion, 'Sin información'),
            'costo':              str(r.Costo) if r.Costo else '',
            'estado':             ESTADOS.get(r.IdEstado, '—'),
            'estado_id':          r.IdEstado or 0,
            'prioridad':          PRIORIDADES.get(r.IdPrioridad, '—'),
            'asignado':           r.NombreUsuariAsig or '',
            'fecha':              _fmt_fecha_hora(r.Fecha)['fecha'],
            'hora':               _fmt_fecha_hora(r.Fecha)['hora'],
            'vencimiento':        _fmt_fecha_hora(r.FechaEstiSoluci)['fecha'],
            'hora_vencimiento':   _fmt_fecha_hora(r.FechaEstiSoluci)['hora'],
            'plan_accion':        r.PlanAccion or '',
            'solucion':           r.Solucion or '',
            'fecha_solucion':     _fmt_fecha_hora(r.FechaRealSoluci)['fecha'],
            'categoria_id':       r.IdCategoria,
            'subcategoria_id':    r.IdSubCategoria,
            'id_usuario_asig':    r.IdUsuarioAsig,
        })

    return _json_ok({'requerimientos': data, 'total': len(data)})


@login_required(login_url='login')
@csrf_exempt
@require_http_methods(['POST'])
def api_req_tic_accion(request, req_id):
    """
    Actualiza plan de acción, reasignación, solución o rechazo de un requerimiento.
    body = {
      accion: 'plan' | 'reasignar' | 'solucionar' | 'rechazar',
      plan_accion: '...',
      id_usuario_asig: 5,       (solo para reasignar)
      solucion: '...',          (solo para solucionar)
      fecha_solucion: 'YYYY-MM-DD'
      motivo: '...',            (solo para rechazar)
    }
    """
    user_id = request.session.get('req_user_id')
    if not user_id:
        return _json_err('Sin permiso', 403)

    try:
        r = Requerimiento.objects.using('requerimientos').get(Codigo=req_id)
    except Requerimiento.DoesNotExist:
        return _json_err('Requerimiento no encontrado', 404)

    body   = json.loads(request.body)
    accion = body.get('accion')

    if accion == 'plan':
        r.PlanAccion = body.get('plan_accion', '').strip()
        r.IdEstado   = 3  # En Proceso
        r.save(using='requerimientos')

    elif accion == 'reasignar':
        nuevo_id = body.get('id_usuario_asig')
        if not nuevo_id:
            return _json_err('id_usuario_asig requerido')
        from requerimientos.models import Usuario as UsuarioReq
        try:
            nuevo = UsuarioReq.objects.using('requerimientos').get(IdUsuario=nuevo_id)
        except UsuarioReq.DoesNotExist:
            return _json_err('Usuario destino no encontrado')
        r.IdUsuarioAsig    = nuevo.IdUsuario
        r.NombreUsuariAsig = nuevo.NombreCompleto
        r.IdEstado         = 2  # Asignado

        categoria_id    = body.get('categoria_id')
        subcategoria_id = body.get('subcategoria_id')
        if categoria_id:
            r.IdCategoria = categoria_id
        if subcategoria_id:
            r.IdSubCategoria = subcategoria_id

        r.save(using='requerimientos')

    elif accion == 'solucionar':
        solucion = body.get('solucion', '').strip()
        fecha    = body.get('fecha_solucion')
        costo    = body.get('costo')
        plan     = body.get('plan_accion')

        if not solucion:
            return _json_err('La solución no puede estar vacía')

        r.Solucion        = solucion
        r.FechaRealSoluci = fecha or None
        r.IdEstado        = 4

        if plan is not None:
            r.PlanAccion = plan.strip()
        if costo not in (None, ''):
            try:
                r.Costo = costo
            except (TypeError, ValueError):
                return _json_err('Costo inválido')

        r.save(using='requerimientos')

    elif accion == 'rechazar':
        # Solo se puede rechazar un requerimiento que ya tiene técnico
        # asignado — si no, al corregirse no habría a quién avisarle.
        if not r.IdUsuarioAsig:
            return _json_err('Este requerimiento no tiene técnico asignado — asígnalo primero antes de poder rechazarlo.')

        motivo = body.get('motivo', '').strip()
        if not motivo:
            return _json_err('El motivo del rechazo no puede estar vacío')

        r.MotivoRechazo = motivo
        r.IdEstado      = 9  # Requiere corrección
        r.save(using='requerimientos')

        # No hay señal (Signals.py) que reaccione a este estado, así que el
        # correo se dispara explícitamente aquí, igual que en el rechazo de
        # aprobación (requerimientos/views.py:_enviar_correo_rechazo).
        from requerimientos.views import _enviar_correo_rechazo_tecnico
        _enviar_correo_rechazo_tecnico(r, motivo)

    else:
        return _json_err('Acción no válida. Use: plan | reasignar | solucionar | rechazar')

    return _json_ok({'codigo': r.codigo(), 'estado_id': r.IdEstado})


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_todos_req_tic(request):
    """Requerimientos SIN ASIGNAR todavía, para la pantalla de Asignar
    Requerimientos — es una cola de pendientes por tomar: en cuanto alguien
    se asigna uno, desaparece de aquí y pasa a su "Mis Requerimientos"."""
    from requerimientos.models import (
        Cargo as CargoReq, TipoRequerimiento, Clasificacion as ClasificacionReq,
        ImagenAdjunta, CentroOperacion as CentroOperacionReq,
    )

    ESTADOS       = {1:'Abierto', 2:'Asignado', 3:'En Proceso', 4:'Cerrado', 5:'Eliminado', 6:'Calificado',
                     7:'Pendiente Aprobación', 8:'Rechazado', 9:'Requiere corrección'}
    PRIORIDADES   = {p.IdPrioridad: p.Descripcion for p in Prioridad.objects.using('requerimientos').all()}
    CATEGORIAS    = {c.IdCategoria: c.Descripcion for c in Categoria.objects.using('requerimientos').all()}
    SUBCATEGORIAS = {s.IdSubCategoria: s.Descripcion for s in SubCategoria.objects.using('requerimientos').all()}
    TIPOS         = {t.IdTipoReque: t.Descripcion for t in TipoRequerimiento.objects.using('requerimientos').all()}
    CARGOS        = {c.IdCargo: c.Descripcion for c in CargoReq.objects.using('requerimientos').all()}
    CLASIFICAC    = {c.IdClasificacion: c.Clasificacion  for c in ClasificacionReq.objects.using('requerimientos').all()}
    # CO en mv_Requerimientos guarda el CÓDIGO corto (ej. "AM1"), no el
    # nombre — se resuelve aquí para que la tabla del técnico sea legible.
    CENTROS       = {c.IdCo: c.Descripcion for c in CentroOperacionReq.objects.using('requerimientos').all()}

    qs = (Requerimiento.objects
          .using('requerimientos')
          .filter(IdUsuarioAsig__isnull=True)
          .exclude(IdEstado__in=[4, 5, 6])
          .order_by('-Fecha'))

    # Adjuntos: solo se guardan los que suba este proyecto (ver ADJUNTO_CARPETA
    # en requerimientos/views.py) — se arma el nombre de archivo en disco
    # igual que en api_adjuntar_archivo: '{IdImagen}_{NombreImagen}'.
    codigos_pagina = [r.Codigo for r in qs]
    ADJUNTOS = {}
    for img in (ImagenAdjunta.objects.using('requerimientos')
                .filter(CodReq__in=codigos_pagina).order_by('IdImagen')):
        nombre_disco = f'{img.IdImagen}_{img.NombreImagen}'
        ADJUNTOS[img.CodReq] = {
            'nombre': img.NombreImagen,
            'url':    f'{settings.MEDIA_URL}requerimientos_adjuntos/{nombre_disco}',
        }

    data = []
    for r in qs:
        adjunto = ADJUNTOS.get(r.Codigo)
        data.append({
            'id':                 r.Codigo,
            'codigo':             r.codigo(),
            'descripcion':        r.Requerimiento or '',
            'fecha':              _fmt_fecha_hora(r.Fecha)['fecha'],
            'solicitante':        r.NombreUsuario or '—',
            'documento':          r.CedulaUsuario or '',
            'correo':             r.Email or '',
            'cargo':              CARGOS.get(r.Cargo, '—'),
            'centro_operacion':   CENTROS.get(r.CO, r.CO or '—'),
            'tipo_requerimiento': TIPOS.get(r.IdTipoReq, '—'),
            'categoria':          CATEGORIAS.get(r.IdCategoria, '—'),
            'subcategoria':       SUBCATEGORIAS.get(r.IdSubCategoria, '—'),
            'prioridad':          PRIORIDADES.get(r.IdPrioridad, '—'),
            'hora':               _fmt_fecha_hora(r.Fecha)['hora'],
            'fecha_vencimiento':  _fmt_fecha_hora(r.FechaEstiSoluci)['fecha'],
            'hora_vencimiento':   _fmt_fecha_hora(r.FechaEstiSoluci)['hora'],
            'asignado':           r.NombreUsuariAsig or '',
            'estado':             ESTADOS.get(r.IdEstado, '—'),
            'clasificacion':      CLASIFICAC.get(r.Clasificacion, 'No hay Clasificación'),
            'categoria_id':       r.IdCategoria,
            'subcategoria_id':    r.IdSubCategoria,
            'plan_accion':        r.PlanAccion or '',
            'costo':              str(r.Costo) if r.Costo else '',
            'archivo_acciones':   '',
            'tiene_adjunto':      bool(adjunto),
            'nombre_adjunto':     adjunto['nombre'] if adjunto else '',
            'url_adjunto':        adjunto['url'] if adjunto else '',
        })
    return _json_ok({'requerimientos': data, 'total': len(data)})


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_notificaciones_bell(request):
    """
    Campanita del header del Dashboard. Combina 4 fuentes, todas calculadas
    en vivo (sin tabla propia de notificaciones):

      1. vencidos               — requerimientos TIC asignados a mí, ya vencidos.
      2. licencias_por_vencer   — licencias de Office que vencen en <=30 días
                                   (o ya vencidas), visible para cualquier usuario
                                   del Dashboard (es info de infraestructura, no
                                   de una persona en particular).
      3. pendientes_aprobacion  — requerimientos en estado "Pendiente Aprobación"
                                   donde YO soy el jefe de área que debe aprobar.
      4. vencidos_sin_asignar   — requerimientos vencidos que nunca se asignaron
                                   a nadie (vista de supervisor). El Dashboard no
                                   tiene control de roles hoy, así que se muestra
                                   igual que el resto de pantallas.

    Distinto del sistema de notificaciones del Portal de Requerimientos
    (mv_Notificaciones / mis_notificaciones): ese avisa al SOLICITANTE de
    que su requerimiento fue asignado/solucionado. Este avisa al TÉCNICO/staff
    del Dashboard de lo que tiene pendiente — son audiencias distintas.
    """
    from datetime import timedelta

    req_user_id = request.session.get('req_user_id')
    hoy = timezone.now().date()

    # Alertas que este usuario ya marcó como leídas (solo aplica a 'licencia'
    # y 'aprobacion' — 'vencido'/'sin_asignar' nunca se filtran aquí).
    leidas = set()
    if req_user_id:
        leidas = set(
            NotificacionBellLeida.objects
            .filter(g235_usuario_id=req_user_id)
            .values_list('g235_tipo', 'g235_referencia_id', 'g235_referencia_fecha')
        )

    # Abierto, Asignado, En Proceso, Pendiente Aprobación — mismos estados
    # "activos" que usa el Portal de Requerimientos para calcular vencidos.
    ESTADOS_ACTIVOS = [1, 2, 3, 7]

    vencidos = []
    pendientes_aprobacion = []
    if req_user_id:
        vencidos_qs = (
            Requerimiento.objects
            .using('requerimientos')
            .filter(IdUsuarioAsig=req_user_id, IdEstado__in=ESTADOS_ACTIVOS, FechaEstiSoluci__lt=hoy)
            .order_by('FechaEstiSoluci')
        )
        vencidos = [{
            'id':             r.Codigo,
            'codigo':         r.codigo(),
            'descripcion':    (r.Requerimiento or '')[:120],
            'fecha_estimada': r.FechaEstiSoluci.strftime('%d/%m/%Y') if r.FechaEstiSoluci else '',
        } for r in vencidos_qs]

        pendientes_aprob_qs = (
            Requerimiento.objects
            .using('requerimientos')
            .filter(IdEstado=7, IdJefeArea=req_user_id)
            .order_by('Fecha')
        )
        for r in pendientes_aprob_qs:
            fecha_str = r.Fecha.strftime('%d/%m/%Y') if r.Fecha else ''
            if ('aprobacion', r.Codigo, fecha_str) in leidas:
                continue
            pendientes_aprobacion.append({
                'id':          r.Codigo,
                'codigo':      r.codigo(),
                'solicitante': r.NombreUsuario or '—',
                'fecha':       fecha_str,
            })

    UMBRAL_DIAS_LICENCIA = 30
    limite_licencia = hoy + timedelta(days=UMBRAL_DIAS_LICENCIA)
    licencias_qs = (
        CaracteristicaLicencia.objects
        .select_related('g227_dispositivo')
        .filter(g227_fecha_vencimiento__isnull=False, g227_fecha_vencimiento__lte=limite_licencia)
        .order_by('g227_fecha_vencimiento')
    )
    licencias_por_vencer = []
    for l in licencias_qs:
        fecha_str = l.g227_fecha_vencimiento.strftime('%d/%m/%Y')
        if ('licencia', l.g227_id, fecha_str) in leidas:
            continue
        licencias_por_vencer.append({
            'id':                 l.g227_id,
            'software':           l.g227_software or 'Licencia',
            'serial_dispositivo': l.g227_dispositivo.g212_serial if l.g227_dispositivo else '—',
            'fecha_vencimiento':  fecha_str,
            'vencida':            l.g227_fecha_vencimiento < hoy,
        })

    sin_asignar_qs = (
        Requerimiento.objects
        .using('requerimientos')
        .filter(IdUsuarioAsig__isnull=True, IdEstado__in=ESTADOS_ACTIVOS, FechaEstiSoluci__lt=hoy)
        .order_by('FechaEstiSoluci')
    )
    vencidos_sin_asignar = [{
        'id':             r.Codigo,
        'codigo':         r.codigo(),
        'descripcion':    (r.Requerimiento or '')[:120],
        'fecha_estimada': r.FechaEstiSoluci.strftime('%d/%m/%Y') if r.FechaEstiSoluci else '',
    } for r in sin_asignar_qs]

    # Recién creados sin asignar — aviso temprano, ANTES de que se venzan.
    # Se excluyen los que ya están vencidos (esos ya salen en
    # vencidos_sin_asignar) para no mostrar el mismo requerimiento dos veces.
    nuevos_qs = (
        Requerimiento.objects
        .using('requerimientos')
        .filter(IdUsuarioAsig__isnull=True, IdEstado=1)
        .filter(Q(FechaEstiSoluci__gte=hoy) | Q(FechaEstiSoluci__isnull=True))
        .order_by('-Fecha')
    )
    nuevos_sin_asignar = []
    for r in nuevos_qs:
        fecha_str = r.Fecha.strftime('%d/%m/%Y') if r.Fecha else ''
        if ('nuevo', r.Codigo, fecha_str) in leidas:
            continue
        nuevos_sin_asignar.append({
            'id':          r.Codigo,
            'codigo':      r.codigo(),
            'descripcion': (r.Requerimiento or '')[:120],
            'solicitante': r.NombreUsuario or '—',
            'fecha':       fecha_str,
        })

    # Préstamos de equipos hechos en el Portal (autoservicio) — avisa que
    # se realizó uno, mientras siga activo (no devuelto) y no se haya
    # marcado como leído. No es una alerta de "vencido": solo informa que
    # ocurrió el préstamo.
    from requerimientos.models import HistorialPrestamo
    prestamos_qs = (
        HistorialPrestamo.objects
        .using('requerimientos')
        .select_related('IdEquipo')
        .filter(FechaDevolucionReal__isnull=True)
        .order_by('-FechaPrestamo')
    )
    prestamos_realizados = []
    for p in prestamos_qs:
        fecha_str = p.FechaPrestamo.strftime('%d/%m/%Y %H:%M') if p.FechaPrestamo else ''
        if ('prestamo', p.IdPrestamo, fecha_str) in leidas:
            continue
        prestamos_realizados.append({
            'id':          p.IdPrestamo,
            'equipo':      p.IdEquipo.NombreEquipo if p.IdEquipo else '—',
            'solicitante': p.NombreSolicitante or '—',
            'area':        p.Area or '',
            'fecha':       fecha_str,
        })

    # Dispositivos de Inventario creados DE HOY EN ADELANTE (fecha fija, no
    # recalculada cada día) que todavía no tienen checklist hecho — solo de
    # los tipos que ya tienen preguntas configuradas (hoy: Portátil). Los
    # dispositivos creados antes de esta fecha quedan fuera a propósito: ya
    # existían antes de que el checklist fuera una función del sistema, así
    # que no cuentan como pendientes "nuevos".
    tipos_con_checklist = ItemChecklist.objects.filter(
        g236_tipo_dispositivo__isnull=False, g236_estado=True
    ).values_list('g236_tipo_dispositivo_id', flat=True).distinct()
    dispositivos_con_checklist = ChecklistDispositivo.objects.filter(
        g237_dispositivo__isnull=False
    ).values_list('g237_dispositivo_id', flat=True).distinct()
    checklist_pendiente_qs = (
        Dispositivo.objects
        .exclude(g212_estado__g201_descripcion__in=ESTADOS_INACTIVOS)
        .filter(g212_tipo_id__in=tipos_con_checklist)
        .filter(g212_fecha_registro__date__gte=CHECKLIST_NOTIF_DESDE)
        .exclude(g212_id__in=dispositivos_con_checklist)
        .select_related('g212_tipo')
        .order_by('-g212_fecha_registro')
    )
    checklist_pendiente = []
    for d in checklist_pendiente_qs:
        fecha_str = d.g212_fecha_registro.strftime('%d/%m/%Y') if d.g212_fecha_registro else ''
        if ('checklist', d.g212_id, fecha_str) in leidas:
            continue
        checklist_pendiente.append({
            'id':     d.g212_id,
            'serial': d.g212_serial,
            'tipo':   d.g212_tipo.g200_tipo_dispositivo if d.g212_tipo else '—',
            'fecha':  fecha_str,
        })

    total_alertas = (
        len(vencidos) + len(licencias_por_vencer)
        + len(pendientes_aprobacion) + len(vencidos_sin_asignar)
        + len(nuevos_sin_asignar) + len(prestamos_realizados)
        + len(checklist_pendiente)
    )

    return _json_ok({
        'vencidos':              vencidos,
        'licencias_por_vencer':  licencias_por_vencer,
        'pendientes_aprobacion': pendientes_aprobacion,
        'vencidos_sin_asignar':  vencidos_sin_asignar,
        'nuevos_sin_asignar':    nuevos_sin_asignar,
        'prestamos_realizados':  prestamos_realizados,
        'checklist_pendiente':   checklist_pendiente,
        'total_alertas':         total_alertas,
    })


@login_required(login_url='login')
@require_http_methods(['POST'])
def api_notificacion_bell_marcar_leida(request):
    """
    Marca como leída una alerta de la campanita (solo 'licencia', 'aprobacion',
    'nuevo', 'prestamo' o 'checklist' — 'vencido'/'sin_asignar' no son marcables, ver api_notificaciones_bell).
    Body: {tipo, referencia_id, referencia_fecha}
    """
    req_user_id = request.session.get('req_user_id')
    if not req_user_id:
        return _json_err('Sesión inválida.')

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_err('JSON inválido')

    tipo             = body.get('tipo')
    referencia_id    = body.get('referencia_id')
    referencia_fecha = body.get('referencia_fecha')

    if tipo not in ('licencia', 'aprobacion', 'nuevo', 'prestamo', 'checklist') or not referencia_id or not referencia_fecha:
        return _json_err('Datos incompletos.')

    NotificacionBellLeida.objects.get_or_create(
        g235_usuario_id=req_user_id,
        g235_tipo=tipo,
        g235_referencia_id=referencia_id,
        g235_referencia_fecha=referencia_fecha,
    )
    return _json_ok({})


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_colaboradores_ti(request):
    """Lista simple de usuarios activos (solo técnicos, TipoUsuario 7 u 8) para el selector de asignación."""
    qs = (Usuario.objects
          .using('requerimientos')
          .filter(Estado=1, TipoUsuario__in=[7, 8])
          .order_by('NombreCompleto'))
    data = [{'id': u.IdUsuario, 'nombre': u.NombreCompleto} for u in qs]
    return _json_ok(data)



def _fmt_fecha_hora(valor):
    """Formatea date/datetime a 'DD/MM/YYYY' y hora 'HH:MM AM/PM' por separado."""
    if not valor:
        return {'fecha': '', 'hora': ''}
    if hasattr(valor, 'hour'):  # es datetime, no solo date
        return {'fecha': valor.strftime('%d/%m/%Y'), 'hora': valor.strftime('%I:%M %p')}
    return {'fecha': valor.strftime('%d/%m/%Y'), 'hora': ''}


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_historial_req_tic(request):
    """
    Todos los requerimientos, sin importar el estado (incluye Cerrado
    y Calificado), para la pantalla de Historial de Requerimientos.
    """
    from requerimientos.models import Clasificacion as ClasificacionReq

    ESTADOS = {1: 'PENDIENTE', 2: 'ASIGNADO', 3: 'EN PROCESO', 4: 'CERRADO', 5: 'ELIMINADO', 6: 'CALIFICADO'}
    PRIORIDADES = {1: 'ALTA', 2: 'MEDIA', 3: 'BAJA'}
    CLASIFICAC  = {c.IdClasificacion: c.Clasificacion for c in ClasificacionReq.objects.using('requerimientos').all()}

    qs = (Requerimiento.objects
          .using('requerimientos')
          .exclude(IdEstado=5)   # oculta solo los eliminados
          .order_by('-Fecha'))

    data = []
    for r in qs:
        estado_id = r.IdEstado or 0
        data.append({
            'id':                  r.Codigo,
            'consecutivo':         r.codigo(),
            'fecha_requerimiento': _fmt_fecha_hora(r.Fecha)['fecha'],
            'remitente':           r.NombreUsuario or '—',
            'descripcion':         r.Requerimiento or '',
            'prioridad':           PRIORIDADES.get(r.IdPrioridad, ''),
            'asignado':            r.NombreUsuariAsig or '',
            'clasificacion':       CLASIFICAC.get(r.Clasificacion, ''),
            'plan_accion':         r.PlanAccion or '',
            'hora_requerimiento':  _fmt_fecha_hora(r.Fecha)['hora'],
            'fecha_solucion':      _fmt_fecha_hora(r.FechaRealSoluci)['fecha'],
            'hora_solucion':       _fmt_fecha_hora(r.FechaRealSoluci)['hora'],
            'solucion':            r.Solucion or '',
            'estado':              ESTADOS.get(estado_id, str(estado_id)),
        })

    return _json_ok({'requerimientos': data, 'total': len(data)})



@login_required(login_url='login')
@require_http_methods(['GET'])
def api_categorias_req(request):
    """Lista de categorías reales (mm_Categoria) para el modal de Asignar."""
    qs = Categoria.objects.using('requerimientos').order_by('Descripcion')
    data = [{'id': c.IdCategoria, 'descripcion': c.Descripcion} for c in qs]
    return _json_ok(data)


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_subcategorias_req(request):
    """Lista de subcategorías (mm_SubCategoria) filtradas por categoria_id."""
    categoria_id = request.GET.get('categoria_id')
    qs = SubCategoria.objects.using('requerimientos').order_by('Descripcion')
    if categoria_id:
        qs = qs.filter(IdCategoria=categoria_id)
    data = [{'id': s.IdSubCategoria, 'descripcion': s.Descripcion, 'categoria_id': s.IdCategoria} for s in qs]
    return _json_ok(data)



# INDICADORES — Panel de requerimientos

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_indicadores_resumen(request):
    """
    Tarjetas de resumen: asignados / sin asignar / en proceso / finalizados.
    Cuenta sobre TODOS los requerimientos vigentes (excluye eliminados).
    """
    qs = Requerimiento.objects.using('requerimientos').exclude(IdEstado=5)

    data = {
        'asignados':   qs.filter(IdEstado=2).count(),
        'sin_asignar': qs.filter(IdEstado=1).count(),
        'en_proceso':  qs.filter(IdEstado=3).count(),
        'finalizados': qs.filter(IdEstado__in=[4, 6]).count(),
    }
    return _json_ok(data)


@login_required(login_url='login')
@require_http_methods(['GET'])
def api_indicadores_tendencia(request):
    """
    Serie diaria de requerimientos por estado (Abiertos / Asignado / En Proceso / Cerrados)
    dentro de un rango de días, con filtro opcional de categoría y subcategoría.
    También calcula el % de cumplimiento (solucionados a tiempo) del rango.

      ?dias=            30 | 15 | 60 | 90   (default 30)
      ?categoria_id=    id de mm_Categoria (opcional)
      ?subcategoria_id= id de mm_SubCategoria (opcional)
    """
    from datetime import date, timedelta

    try:
        dias = int(request.GET.get('dias', 30))
    except (TypeError, ValueError):
        dias = 30
    dias = max(1, min(dias, 365))

    categoria_id    = request.GET.get('categoria_id') or None
    subcategoria_id = request.GET.get('subcategoria_id') or None

    fecha_fin   = date.today()
    fecha_inicio = fecha_fin - timedelta(days=dias)

    qs = (Requerimiento.objects
          .using('requerimientos')
          .exclude(IdEstado=5)
          .filter(Fecha__gte=fecha_inicio, Fecha__lte=fecha_fin))

    if categoria_id:
        qs = qs.filter(IdCategoria=categoria_id)
    if subcategoria_id:
        qs = qs.filter(IdSubCategoria=subcategoria_id)

    # Construir un diccionario fecha -> conteos por estado
    dias_map = {}
    d = fecha_inicio
    while d <= fecha_fin:
        dias_map[d] = {'abiertos': 0, 'asignado': 0, 'en_proceso': 0, 'cerrados': 0}
        d += timedelta(days=1)

    for r in qs:
        if not r.Fecha or r.Fecha not in dias_map:
            continue
        if r.IdEstado == 1:
            dias_map[r.Fecha]['abiertos'] += 1
        elif r.IdEstado == 2:
            dias_map[r.Fecha]['asignado'] += 1
        elif r.IdEstado == 3:
            dias_map[r.Fecha]['en_proceso'] += 1
        elif r.IdEstado in (4, 6):
            dias_map[r.Fecha]['cerrados'] += 1

    serie = [
        {
            'fecha':      f.strftime('%Y-%m-%d'),
            'abiertos':   v['abiertos'],
            'asignado':   v['asignado'],
            'en_proceso': v['en_proceso'],
            'cerrados':   v['cerrados'],
        }
        for f, v in sorted(dias_map.items())
    ]

    # % de cumplimiento: solucionados dentro del rango, a tiempo vs. total solucionados
    cerrados_qs = qs.filter(IdEstado__in=[4, 6])
    total_cerrados = cerrados_qs.count()
    a_tiempo = cerrados_qs.filter(
        FechaRealSoluci__isnull=False,
        FechaEstiSoluci__isnull=False,
        FechaRealSoluci__lte=F('FechaEstiSoluci'),
    ).count()
    pct_cumplimiento = round((a_tiempo / total_cerrados) * 100, 1) if total_cerrados else 0

    return _json_ok({
        'serie':            serie,
        'pct_cumplimiento': pct_cumplimiento,
        'total_cerrados':   total_cerrados,
        'a_tiempo':         a_tiempo,
    })
    

@login_required(login_url='login')
@require_http_methods(['GET'])
def api_indicadores_calificacion(request):
    """
    Calificación de calidad (mv_EvaluacionReq): satisfacción real del usuario
    que reportó el requerimiento, NO la puntualidad del técnico.

      ?dias=            30 | 15 | 60 | 90   (default 30)
      ?categoria_id=    id de mm_Categoria (opcional)
      ?subcategoria_id= id de mm_SubCategoria (opcional)

    Responde:
      promedio             — promedio general (1 a 5) en el rango/filtro
      total_evaluaciones   — cuántas evaluaciones entran en el filtro
      distribucion         — {'1': n, '2': n, '3': n, '4': n, '5': n}
      tendencia            — [{semana, promedio, cantidad}, ...] por semana
    """
    from datetime import date, timedelta
    from collections import defaultdict
    from requerimientos.models import EvaluacionReq

    try:
        dias = int(request.GET.get('dias', 30))
    except (TypeError, ValueError):
        dias = 30
    dias = max(1, min(dias, 365))

    categoria_id    = request.GET.get('categoria_id') or None
    subcategoria_id = request.GET.get('subcategoria_id') or None

    fecha_fin    = date.today()
    fecha_inicio = fecha_fin - timedelta(days=dias)

    # 1. Requerimientos que caen en el filtro (fecha + categoría/subcategoría)
    req_qs = (Requerimiento.objects
              .using('requerimientos')
              .exclude(IdEstado=5)
              .filter(Fecha__gte=fecha_inicio, Fecha__lte=fecha_fin))
    if categoria_id:
        req_qs = req_qs.filter(IdCategoria=categoria_id)
    if subcategoria_id:
        req_qs = req_qs.filter(IdSubCategoria=subcategoria_id)

    codigos_fecha = {r.Codigo: r.Fecha for r in req_qs.only('Codigo', 'Fecha')}
    codigos = list(codigos_fecha.keys())

    # 2. Evaluaciones de esos requerimientos
    evals = list(
        EvaluacionReq.objects
        .using('requerimientos')
        .filter(IdReq__in=codigos, Evaluacion__isnull=False)
        .values('IdReq', 'Evaluacion')
    )

    valores = [e['Evaluacion'] for e in evals if 1 <= (e['Evaluacion'] or 0) <= 5]
    total   = len(valores)
    promedio = round(sum(valores) / total, 2) if total else 0

    distribucion = {str(n): 0 for n in range(1, 6)}
    for v in valores:
        distribucion[str(v)] += 1

    # 3. Tendencia semanal del promedio (agrupado por semana de la fecha del requerimiento)
    semana_map = defaultdict(list)
    for e in evals:
        val = e['Evaluacion']
        if not (1 <= (val or 0) <= 5):
            continue
        fecha_req = codigos_fecha.get(e['IdReq'])
        if not fecha_req:
            continue
        inicio_semana = fecha_req - timedelta(days=fecha_req.weekday())
        semana_map[inicio_semana].append(val)

    tendencia = [
        {
            'semana':   f.strftime('%d/%m'),
            'promedio': round(sum(vs) / len(vs), 2),
            'cantidad': len(vs),
        }
        for f, vs in sorted(semana_map.items())
    ]

    return _json_ok({
        'promedio':           promedio,
        'total_evaluaciones': total,
        'distribucion':       distribucion,
        'tendencia':          tendencia,
    })